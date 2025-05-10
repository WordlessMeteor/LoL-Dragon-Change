from lcu_driver import Connector
import argparse, copy, os, pandas, psutil, re, requests, shutil, subprocess, time, traceback, unicodedata, uuid
from urllib.parse import quote, unquote
from wcwidth import wcswidth
from openpyxl.styles import Color, numbers, PatternFill
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter

parser = argparse.ArgumentParser()
parser.add_argument("-l", "--locale", help = "选择非《英雄联盟》本地化内容的显示语言（Choose the display language that isn't part of League of Legends localization content）", action = "store")
parser.add_argument("-os", "--open_after_save", help = "在成功保存文件后，跳过程序询问，直接打开该文件（Once the workbook is saved successfully, directly open this workbook without program asking）", action = "store_true")
parser.add_argument("--verbose", help = "显示详细加载进度（Display detailed loading process）", action = "store_true")
parser.add_argument("--nonverbose", help = "显示简略加载进度（Display simplified loading process）", action = "store_true")
args = parser.parse_args()
if args.locale == "en_US":
    arg_locale = "en_US" #只影响常量字典的本地化（Only influences the localization of the constant dictionaries）
else:
    arg_locale = "zh_CN"

#=============================================================================
# * 声明（Declaration）
#=============================================================================
# 作者（Author）：       XHXIAIEIN
# 更新（Last update）：  2021/01/08
# 主页（Home page）：    https://github.com/XHXIAIEIN/LeagueCustomLobby/
#=============================================================================

#-----------------------------------------------------------------------------
# 工具库（Tool library）
#-----------------------------------------------------------------------------
#  - lcu-driver 
#    https://github.com/sousa-andre/lcu-driver
#-----------------------------------------------------------------------------

connector = Connector()

async def get_summoner_data(connection):
    data = await connection.request('GET', '/lol-summoner/v1/current-summoner')
    summoner = await data.json()
    print("displayName:    %s" %(summoner["gameName"] + "#" + summoner["tagLine"]))
    print("summonerId:     %s" %(summoner["summonerId"]))
    print("puuid:          %s" %(summoner["puuid"]))
    print("-")


#-----------------------------------------------------------------------------
#  lockfile
#-----------------------------------------------------------------------------
async def update_lockfile(connection):
    path = os.path.join(connection.installation_path.encode('gb18030').decode('utf-8'), 'lockfile')
    if os.path.isfile(path):
        file = open(path, 'w+')
        text = "LeagueClient:%d:%d:%s:%s" %(connection.pid, connection.port, connection.auth_key, connection.protocols[0])
        file.write(text)
        file.close()
    return None

async def get_lockfile(connection):
    path = os.path.join(connection.installation_path.encode('gb18030').decode('utf-8'), 'lockfile')
    if os.path.isfile(path):
        file = open(path, 'r')
        text = file.readline().split(':')
        file.close()
        print(connection.address)
        print(f'riot    {connection.auth_key}')
        return connection.auth_key
    return None

#-----------------------------------------------------------------------------
# 汇总英雄选择阶段或游戏内玩家的战绩（Summarize players' stats in recent matches during champ select stage or in game）
#-----------------------------------------------------------------------------

def count_nonASCII(s: str): #统计一个字符串中占用命令行2个宽度单位的字符个数（Count the number of characters that take up 2 width unit in CMD）
    return sum([unicodedata.east_asian_width(character) in ("F", "W") for character in list(str(s))])

def format_df(df: pandas.DataFrame, width_exceed_ask: bool = True, direct_print: bool = False, print_header: bool = True, print_index: bool = False, reserve_index = False, start_index = 0, header_align: str = "^", align: str = "^", align_replicate_rule: str = "all"): #按照每列最长字符串的命令行宽度加上2，再根据每个数据的中文字符数量决定最终格式化输出的字符串宽度（Get the width of the longest string of each column, add it by 2, and substract it by the number of each cell string's Chinese characters to get the final width for each cell to print using `format` function）
    old_index = df.index
    df.index = range(start_index, len(df) + start_index)
    maxLens = {}
    maxWidth = shutil.get_terminal_size()[0]
    fields = df.columns.tolist()
    for field in fields:
        maxLens[field] = max(0 if len(df) == 0 else max(map(lambda x: wcswidth(str(x)), df[field])), wcswidth(str(field))) + 2
    index_len = max(len(str(start_index)), len(str(start_index + len(df) - 1)))
    if sum(maxLens.values()) + 2 * (len(fields) - 1) > maxWidth or print_index and index_len + sum(maxLens.values()) + 2 * len(fields) > maxWidth:
        if width_exceed_ask:
            print("单行数据字符串输出宽度超过当前终端窗口宽度！是否继续？（输入任意键继续，否则直接打印该数据框。）\nThe output width of each record string exceeds the current width of the terminal window! Continue? (Input anything to continue, or null to directly print this dataframe.)")
            if input() == "":
                #print(df)
                result = str(df)
                return (result, maxLens)
        elif direct_print:
            print("单行数据字符串输出宽度超过当前终端窗口宽度！将直接打印该数据框！\nThe output width of each record string exceeds the current width of the terminal window! The program is going to directly print this dataframe!")
            result = str(df)
            return (result, maxLens)
        else:
            print("单行数据字符串输出宽度超过当前终端窗口宽度！将继续格式化输出！\nThe output width of each record string exceeds the current width of the terminal window! The program is going on formatted printing!")
    result = ""
    #确定各列的排列方向（Determine the alignments of all columns）
    if isinstance(header_align, str) and isinstance(align, str):
        if not all(map(lambda x: x in {"<", "^", ">"}, header_align)) or not all(map(lambda x: x in {"<", "^", ">"}, align)):
            print('排列方式字符串参数错误！排列方式必须是“<”“^”或者“>”中的一个。请修改排列方式字符串参数。\nParameter ERROR of the alignment string! The alignment value must be one of {"<", "^", ">"}. Please change the alignment string parameter.')
        if len(header_align) == 0: #指定为空字符串，即默认居中输出（Specifying it as a null string means output centered by default）
            header_alignments = ["^"] * df.shape[1]
        elif len(header_align) == 1:
            header_alignments = [header_align] * df.shape[1]
        else:
            header_alignments_tmp = list(header_align)
            if len(header_align) < df.shape[1]:
                if align_replicate_rule == "last":
                    header_alignments = header_alignments_tmp + [header_alignments_tmp[-1]] * len(df.shape[1] - len(header_align))
                else:
                    if align_replicate_rule != "all":
                        print("排列方式列表补充规则不合法！将默认采用全部填充。\nAlignment list supplement rule illegal! The whole alignment string will be replicated.")
                    header_alignments = header_alignments_tmp * (df.shape[1] // len(header_align)) + header_alignments_tmp[:df.shape[1] % len(header_align)]
            else:
                header_alignments = header_alignments_tmp[:df.shape[1]]
        if len(align) == 0:
            alignments = ["^"] * df.shape[1]
        elif len(align) == 1:
            alignments = [align] * df.shape[1]
        else:
            alignments_tmp = list(align)
            if len(align) < df.shape[1]:
                if align_replicate_rule == "last":
                    alignments = alignments_tmp + [alignments_tmp[-1]] * len(df.shape[1] - len(align))
                else:
                    if align_replicate_rule != "all":
                        print("排列方式列表补充规则不合法！将默认采用全部填充。\nAlignment list supplement rule illegal! The whole alignment string will be replicated.")
                    alignments = alignments_tmp * (df.shape[1] // len(align)) + alignments_tmp[:df.shape[1] % len(align)]
            else:
                alignments = alignments_tmp[:df.shape[1]]
        if print_header:
            if print_index:
                result += " " * (index_len + 2)
            for i in range(df.shape[1]):
                field = fields[i]
                tmp = "{0:{align}{w}}".format(field, align = header_alignments[i], w = maxLens[str(field)] - count_nonASCII(str(field)))
                result += tmp
                #print(tmp, end = "")
                if i != df.shape[1] - 1:
                    result += "  "
                    #print("  ", end = "")
            result += "\n"
            #print()
        index = start_index
        for i in range(df.shape[0]):
            if print_index:
                result += "{0:>{w}}".format(old_index[index - start_index] if reserve_index else index, w = index_len) + "  "
            for j in range(df.shape[1]):
                field = fields[j]
                cell = str(list(df[field])[i])
                tmp = "{0:{align}{w}}".format(cell, align = alignments[j], w = maxLens[field] - count_nonASCII(cell))
                result += tmp
                #print(tmp, end = "")
                if j != df.shape[1] - 1:
                    result += "  "
                    #print("  ", end = "")
            if i != df.shape[0] - 1:
                result += "\n"
            #print() #注意这里的缩进和上一行不同（Note that here the indentation is different from the last line）
            index += 1
    else:
        print("排列方式参数错误！请传入字符串。\nAlignment parameter ERROR! Please pass a string instead.")
    return (result, maxLens)

def lcuTimestamp(timestamp): #根据对局时间轴的时间戳返回对局时间（Return the time according to the timestamp in match timeline）
    min = timestamp // 60
    sec = timestamp % 60
    return str(min) + ":" + "{0:0>2}".format(str(sec))

def patch_compare(patch1, patch2): #比较两个版本号的先后顺序。当patch1 < patch2时，返回True，否则返回False。用于比较DataDragon数据库中未收录的版本和收录的最新版本的关系。如果未收录的版本小于收录的最新版本，那么该版本是美测服的临时版本，后来被合并更新了，如正式服将13.2和13.3合并更新了，因此DataDragon数据库中未收录13.2版本的数据；如果未收录的版本大于收录的最新版本，那么该版本是美测服的当前版本，但是仍处于开发状态，尚未完全确定，所以DataDragon数据库尚未收录，将以最新版本代替该版本；二者不可能相等，因为如果相等的话，就不会引发报错而调用此函数（Compare the time order of two patches. When patch1 < patch2, return True and vice versa. Designed to compare a patch not archived in DataDragon database with the latest patch archived in DataDragon database. If the unarchived patch is less than the latest archived patch, then this patch must be the intermediate patch and be merged into the update of its successive patch, such as Patch 13.2 merged into the update of Patch 13.3, so that DataDragon database doesn't archive the data of Patch 13.2; If the unarchived patch is greater than the latest archived patch, then this patch must be the current patch on PBE but is under development and improvement, so that DataDragon database doesn't archive this patch, either, in which case the latest patch will be used to substitute this unarchived patch; The two patches can't be the same, for suppose they're same, then the error to cause the call of this function won't be triggered）
    if not isinstance(patch1, str):
        patch1 = str(patch1)
    if not isinstance(patch2, str):
        patch2 = str(patch2)
    lst1, lst2 = patch1.split("."), patch2.split(".")
    try:
        lst1 = list(map(int, lst1))
    except ValueError:
        print("第1个版本字符串不合法！请输入用半角句号连接的正整数，如13.15.1、10.10.3216176。\nThe first patch variable is illegal! Please pass the integers concatenated by dot, such as 13.15.1 and 10.10.3216176.")
        return 1
    try:
        lst2 = list(map(int, lst2))
    except ValueError:
        print("第2个版本字符串不合法！请输入用半角句号连接的正整数，如13.15.1、10.10.3216176。\nThe second patch variable is illegal! Please pass the integers concatenated by dot, such as 13.15.1 and 10.10.3216176.")
        return 1
    for i in range(min(len(lst1), len(lst2))):
        if lst1[i] < lst2[i]:
            return True
        elif lst1[i] > lst2[i]:
            return False
        else:
            continue
    if len(lst1) < len(lst2):
        return True
    else:
        return False #这里将两个版本相同视为假，暗示了在本程序用得到的地方，两个版本不可能相同（Here the case where the two patches are the same is regarded as False, which indicates that the two patches can't be same within its use in this program）

def FindPostPatch(patch, patchList): #二分查找某个版本号在DataDragon数据库的后一个版本（Binary search for the precedent patch of a given patch in the patch list archived in DataDragon database）
    leftIndex, rightIndex = 0, len(patchList) - 1
    mid = (leftIndex + rightIndex) // 2
    count = 0 #函数调试阶段的保护机制（A protecion mechanism during rebugging this function）
    #print("[" + str(count) + "]", leftIndex, mid, rightIndex)
    while leftIndex < rightIndex:
        count += 1
        if patch_compare(patch, patchList[mid]):
            leftIndex = mid + 1
            mid = (leftIndex + rightIndex) // 2
        elif patch_compare(patchList[mid], patch):
            rightIndex = mid
            mid = (leftIndex + rightIndex) // 2
        else:
            return patchList[mid - 1]
        #print("[" + str(count) + "]", leftIndex, mid, rightIndex)
        if count >= 15:
            print("程序即将进入死循环！请检查算法！\nThe program is stepping into a dead loop! Please check the algorithm!")
            return 1
    if mid >= 1:
        return patchList[mid - 1]
    else:
        print("该版本为美测服最新版本，暂未收录在DataDragon数据库中。\nThis version is the latest version on PBE and isn't archived in DataDragon database for now.")
        return "pbe"

def subscope(scope: dict = {}):
    s = copy.deepcopy(scope)
    while True:
        expr = input()
        tokens = expr.split() #去除空格的词法分析（Parse by spliting by space）
        if expr == "-1":
            break
        elif expr == "0":
            s = copy.deepcopy(scope)
            print("变量和作用域已复位。\nVariables and the scope have been reset.")
        else:
            try:
                exec(expr, s)
            except:
                traceback_info = traceback.format_exc()
                print(traceback_info)
    return 0

def check_proc_trees(pids: list[int]) -> pandas.DataFrame: #查看本程序创建的所有进程及其子进程（Check all processes and subprocess created by this program）
    process_header = {"No.": "序号", "pid": "进程号", "name": "名称", "createTime": "进程创建时间", "status": "状态"}
    process_header_keys = list(process_header.keys())
    process_data = {}
    for i in range(len(process_header_keys)):
        key = process_header_keys[i]
        process_data[key] = []
    for i in range(len(pids)):
        pid = pids[i]
        count = 0
        try:
            parent = psutil.Process(pid)
        except psutil.NoSuchProcess:
            pass
        else:
            count += 1
            for j in range(len(process_header_keys)):
                key = process_header_keys[j]
                if j == 0:
                    process_data[key].append("%d.%d" %(i + 1, count))
                elif j == 3:
                    process_data[key].append(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(parent.create_time())))
                else:
                    process_data[key].append(eval(f"parent.{key}"))
            for child in parent.children(recursive = True):
                count += 1
                for j in range(len(process_header_keys)):
                    key = process_header_keys[j]
                    if j == 0:
                        process_data[key].append("%d.%d" %(i + 1, count))
                    elif j == 3:
                        process_data[key].append(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(child.create_time())))
                    else:
                        process_data[key].append(eval(f"child.{key}"))
    process_df = pandas.DataFrame(data = process_data)
    process_df = pandas.concat([pandas.DataFrame([process_header])[process_df.columns], process_df], ignore_index = True)
    return process_df

def kill_proc_tree(pid: int, including_parent: bool = True): #清理残留进程及其子进程（Clear the remaining processes）
    parent = psutil.Process(pid)
    for child in parent.children(recursive = True):
        child.kill()
    if including_parent:
        parent.kill()

async def get_info(connection, name: str, searchType: str | int = "riotId"):
    #searchTypes = {0: "selfCheck", 1: "riotId", 2: "puuid", 3: "summonerId"}
    current_info = await (await connection.request("GET", "/lol-summoner/v1/current-summoner")).json()
    result = {"searchType": "riotId", "endpoint": "/lol-summoner/v2/summoners/puuid/{puuid}", "info_got": False, "network_error": False, "body": {}, "message": "", "selfInfo": False}
    try:
        name = int(name)
    except ValueError:
        if name == "current-summoner":
            result = {"searchType": "selfCheck", "endpoint": "/lol-summoner/v1/current-summoner", "info_got": True, "network_error": False, "body": current_info, "message": "", "selfInfo": True}
        elif name.count("-") == 4 and len(name.replace(" ", "")) > 22: #拳头规定的玩家昵称不超过16个字符，昵称编号不超过5个字符（Riot game name can't exceed 16 characters. The tagline can't exceed 5 characters）
            result["searchType"] = "puuid"
            result["endpoint"] = "/lol-summoner/v2/summoners/puuid/{puuid}"
            info = await (await connection.request("GET", f"/lol-summoner/v2/summoners/puuid/{name}")).json()
            result["body"] = info
            if "errorCode" in info:
                if info["httpStatus"] == 400:
                    result["message"] = "您输入的玩家通用唯一识别码格式有误！请重新输入！\nPUUID wasn't in UUID format! Please try again!"
                elif info["httpStatus"] == 404:
                    result["message"] = "未找到玩家通用唯一识别码为%s的玩家；请核对识别码并稍后再试。\nA player with puuid %s was not found; verify the puuid and try again." %(name, name)
                else:
                    result["network_error"] = True
                    result["message"] = "网络异常。\nNetwork Error."
            else:
                result["info_got"] = True
                result["selfInfo"] = info["puuid"] == current_info["puuid"]
        else:
            result["searchType"] = "riotId"
            result["endpoint"] = "/lol-summoner/v1/summoners?name={name}"
            if name.count("#") == 0:
                result["message"] = '召唤师名称已变更为拳头ID。请以“{玩家昵称}#{昵称编号}”的格式输入。\nSummoner name has been replaced with Riot ID. Please input the name in this format: "{gameName}#{tagLine}", e.g. "%s#%s".' %(current_info["gameName"], current_info["tagLine"])
            elif name.count("#") > 1:
                result["message"] = "该玩家名字包含了无效字符。\nThis player name contains invalid characters."
            else:
                gameName, tagLine = name.split("#")
                if len(gameName) == 0:
                    result["message"] = "缺少玩家昵称。\nGame name is missing."
                elif len(tagLine) == 0:
                    result["message"] = "缺少昵称编号。\nTagline is missing."
                elif len(gameName) < 3:
                    result["message"] = "召唤师昵称过短。\nRiot ID is too short."
                elif len(gameName.replace(" ", "")) > 16:
                    result["message"] = "召唤师昵称过长。\nRiot ID is too long."
                else:
                    info = await (await connection.request("GET", "/lol-summoner/v1/summoners?name=" + quote(name))).json()
                    result["body"] = info
                    if "errorCode" in info:
                        if info["httpStatus"] == 404:
                            result["message"] = "未找到%s；请核对下名字并稍后再试。\n%s was not found; verify the name and try again." %(name, name)
                        else:
                            result["network_error"] = True
                            result["message"] = "网络异常。\nNetwork Error."
                    else:
                        result["info_got"] = True
                        result["selfInfo"] = info["puuid"] == current_info["puuid"]
    else:
        result["searchType"] = "summonerId"
        result["endpoint"] = "/lol-summoner/v1/summoners/{id}"
        info = await (await connection.request("GET", f"/lol-summoner/v1/summoners/{name}")).json()
        result["body"] = info
        if "errorCode" in info:
            if info["httpStatus"] == 400:
                if info["message"] == "Value %d for 'id' of type uint64 is out of range":
                    result["message"] = "您输入的召唤师序号格式有误！请重新输入！\nValue for 'id' of type uint64 is out of range! Please try again!"
                else:
                    result["message"] = "未找到召唤师序号为%s的玩家；请核对召唤师序号并稍后再试。\nA player with puuid %s was not found; verify the summonerId and try again." %(name, name)
            elif info["httpStatus"] == 404:
                result["message"] = "未找到召唤师序号为%s的玩家；请核对召唤师序号并稍后再试。\nA player with puuid %s was not found; verify the summonerId and try again." %(name, name)
            else:
                result["network_error"] = True
                result["message"] = "网络异常。\nNetwork Error."
        else:
            result["info_got"] = True
            result["selfInfo"] = info["puuid"] == current_info["puuid"]
    return result

def get_info_name(info: dict, mode = 1) -> str:
    if not isinstance(info, dict) or not all(i in info for i in ["displayName", "gameName", "tagLine"]):
        print("您的召唤师信息格式有误！\nERROR format of summoner information!")
        name = ""
        exit()
    else:
        if info["displayName"] or info["gameName"]:
            if info["gameName"] and info["tagLine"]:
                name = info["gameName"] + "#" + info["tagLine"]
            elif not info["tagLine"] and info["gameName"]:
                name = info["gameName"]
            else:
                name = info["displayName"]
        else: #新玩家属于这种类型（This case matches new players）
            if mode == 1:
                name = str(info["puuid"])
            elif mode == 2: #仅用于设置召唤师数据保存路径（Designed to set the summoner name directory）
                name = "0. 新玩家\\" + str(info["puuid"])
            elif mode == 3: #仅用于设置召唤师数据保存路径（Designed to set the summoner name directory）
                name = "0. New Player\\" + str(info["puuid"])
    return name

def verify_uuid(s: str) -> bool:
    try:
        return s == str(uuid.UUID(s))
    except ValueError:
        return False

async def define_global_variables(connection) -> None:
    #准备数据资源（Prepare data resources）
    print("正在准备数据资源……\nPreparing data resources ...")
    global gamemodes, spells, LoLChampions, LoLItems, summonerIcons, perks, perkstyles, TFTAugments, TFTChampions, TFTItems, TFTCompanions, TFTTraits, CherryAugments
    ##游戏模式（Game mode）
    if not args.nonverbose:
        print("正在加载游戏模式信息……\nLoading game mode information ...")
    gamemode = await (await connection.request("GET", "/lol-game-queues/v1/queues")).json()
    gamemodes = {-1: {"name": "自定义", "gameMode": "CUSTOM", "category": "CUSTOM", "description": "", "type": "CUSTOM"}, 0: {"name": "自定义", "gameMode": "CUSTOM", "category": "CUSTOM", "description": "", "type": "CUSTOM"}}
    for gamemode_iter in gamemode:
        gamemode_id = gamemode_iter["id"]
        gamemodes_iter = {}
        gamemodes_iter["name"] = gamemode_iter["name"]
        gamemodes_iter["gameMode"] = gamemode_iter["gameMode"]
        gamemodes_iter["category"] = gamemode_iter["category"]
        gamemodes_iter["description"] = gamemode_iter["description"]
        gamemodes_iter["type"] = gamemode_iter["type"]
        gamemodes[gamemode_id] = gamemodes_iter
    ##召唤师技能（Summoner spell）
    if not args.nonverbose:
        print("正在加载召唤师技能信息……\nLoading summoner spell information ...")
    spell_initial = await (await connection.request("GET", "/lol-game-data/assets/v1/summoner-spells.json")).json()
    spells = {}
    for spell_iter in spell_initial:
        spell_id = spell_iter["id"]
        spells[spell_id] = spell_iter
    ##英雄（Champion）
    if not args.nonverbose:
        print("正在加载英雄信息……\nLoading champion information ...")
    LoLChampion_initial = await (await connection.request("GET", "/lol-game-data/assets/v1/champion-summary.json")).json()
    LoLChampions = {}
    for LoLChampion_iter in LoLChampion_initial:
        LoLChampion_id = LoLChampion_iter["id"]
        LoLChampions[LoLChampion_id] = LoLChampion_iter
    ##英雄联盟装备（LoL item）
    if not args.nonverbose:
        print("正在加载英雄联盟装备信息……\nLoading LoL item information ...")
    LoLItem_initial = await (await connection.request("GET", "/lol-game-data/assets/v1/items.json")).json()
    LoLItems = {}
    for LoLItem_iter in LoLItem_initial:
        LoLItem_id = int(LoLItem_iter["id"])
        LoLItems[LoLItem_id] = LoLItem_iter
    ##召唤师图标（Summoner icon）
    if not args.nonverbose:
        print("正在加载召唤师图标信息……\nLoading summoner icon information ...")
    summonerIcon_initial = await (await connection.request("GET", "/lol-game-data/assets/v1/summoner-icons.json")).json()
    summonerIcons = {}
    for summonerIcon_iter in summonerIcon_initial:
        summonerIcon_id = int(summonerIcon_iter["id"])
        summonerIcons[summonerIcon_id] = summonerIcon_iter
    ##符文（Perk）
    if not args.nonverbose:
        print("正在加载基石符文信息……\nLoading perk information ...")
    perk_initial = await (await connection.request("GET", "/lol-game-data/assets/v1/perks.json")).json()
    perks = {}
    for perk_iter in perk_initial:
        perk_id = perk_iter["id"]
        perks[perk_id] = perk_iter
    ##符文系（Perkstyle）
    if not args.nonverbose:
        print("正在加载符文系信息……\nLoading perkstyle information ...")
    perkstyle_initial = await (await connection.request("GET", "/lol-game-data/assets/v1/perkstyles.json")).json()
    perkstyles = {}
    for perkstyle_iter in perkstyle_initial["styles"]:
        perkstyle_id = perkstyle_iter["id"]
        perkstyles[perkstyle_id] = perkstyle_iter
    ##云顶之弈强化符文（TFT augments）
    if not args.nonverbose:
        print("正在加载云顶之弈基础数据……\nLoading TFT basic data from CommunityDragon ...")
    platform_config = await (await connection.request("GET", "/lol-platform-config/v1/namespaces")).json()
    platformId = platform_config["LoginDataPacket"]["platformId"]
    URLPatch = "pbe" if platformId == "PBE1" or platformId == "PBE" else "latest"
    region_locale = await (await connection.request("GET", "/riotclient/region-locale")).json()
    locale = region_locale["locale"].lower()
    TFT_initial = requests.get(f"https://raw.communitydragon.org/{URLPatch}/cdragon/tft/{locale}.json").json()
    TFTAugments = {}
    for item in TFT_initial["items"]:
        item_apiName = item["apiName"]
        TFTAugments[item_apiName] = item
    ##云顶之弈英雄（TFT champion）
    if not args.nonverbose:
        print("正在加载云顶之弈棋子信息……\nLoading TFT champion information ...")
    TFTChampion_initial = await (await connection.request("GET", "/lol-game-data/assets/v1/tftchampions.json")).json()
    TFTChampions = {}
    for TFTChampion_iter in TFTChampion_initial:
        champion_name = TFTChampion_iter["name"]
        TFTChampions[champion_name] = TFTChampion_iter["character_record"]
    ##云顶之弈装备（TFT item）
    if not args.nonverbose:
        print("正在加载云顶之弈装备信息……\nLoading TFT item information ...")
    TFTItem_initial = await (await connection.request("GET", "/lol-game-data/assets/v1/tftitems.json")).json()
    TFTItems = {}
    for TFTItem_iter in TFTItem_initial:
        item_nameId = TFTItem_iter["nameId"]
        TFTItems[item_nameId] = TFTItem_iter
    ##云顶之弈小小英雄（TFT companion）
    if not args.nonverbose:
        print("正在加载云顶之弈小小英雄信息……\nLoading companion information ...")
    TFTCompanion_initial = await (await connection.request("GET", "/lol-game-data/assets/v1/companions.json")).json()
    TFTCompanions = {}
    for companion_iter in TFTCompanion_initial:
        contentId = companion_iter["contentId"]
        TFTCompanions[contentId] = companion_iter
    ##云顶之弈羁绊（TFT Trait）
    if not args.nonverbose:
        print("正在加载云顶之弈羁绊信息……\nLoading TFT trait information ...")
    TFTTrait_initial = await (await connection.request("GET", "/lol-game-data/assets/v1/tfttraits.json")).json()
    TFTTraits = {}
    for trait_iter in TFTTrait_initial:
        trait_id = trait_iter["trait_id"]
        conditional_trait_sets = {}
        for conditional_trait_set in trait_iter["conditional_trait_sets"]:
            style_idx = conditional_trait_set["style_idx"]
            conditional_trait_sets[style_idx] = conditional_trait_set
        trait_iter["conditional_trait_sets"] = conditional_trait_sets
        TFTTraits[trait_id] = trait_iter
    ##斗魂竞技场强化符文（Arena augment）
    if not args.nonverbose:
        print("正在加载斗魂竞技场强化符文信息……\nLoading Arena augment information ...")
    CherryAugment_initial = await (await connection.request("GET", "/lol-game-data/assets/v1/cherry-augments.json")).json()
    CherryAugments = {}
    for CherryAugment in CherryAugment_initial:
        CherryAugment_id = CherryAugment["id"]
        CherryAugments[CherryAugment_id] = CherryAugment
    #定义常量字典（Define constant dictionaries）
    print("初始化常量字典……\nInitializing constant dictionaries ...")
    global categories, gameSelectCategories, gameSelectModeGroups, queueAvailability_dict, queueTypes, banModes, pickModes, tiers, ratedTiers, tiers_all, gameTypes, team_color, subteam_color, endOfGameResults, augment_rarity, lanes, roles, traitStyles, rarities, win_dict
    if arg_locale == "zh_CN":
        categories = {"Custom": "自定义对局", "PvP": "玩家对战", "VersusAi": "人机对战"}
        gameSelectCategories = {"": "待定", "CreateCustom": "创建自定义对局", "JoinCustom": "加入自定义对局", "kPvP": "玩家对战", "kTraining": "训练", "kVersusAI": "人机对战"}
        gameSelectModeGroups = {"": "待定", "kARAM": "极地大乱斗", "kAlternativeLeagueGameModes": "轮换《英雄联盟》游戏模式", "kSummonersRift": "召唤师峡谷", "kTeamfightTactics": "云顶之弈"}
        queueAvailability_dict = {"Available": "√", "PlatformDisabled": ""}
        queueTypes = {"RANKED_SOLO_5x5": "单人/双人", "RANKED_FLEX_SR": "灵活 5V5", "RANKED_TFT": "云顶之弈", "RANKED_TFT_PAIRS": "2V0", "RANKED_TFT_DOUBLE_UP": "双人作战", "RANKED_TFT_TURBO": "狂暴模式", "CHERRY": "斗魂竞技场"}
        banModes = {"": "待定", "SkipBanStrategy": "无", "StandardBanStrategy": "经典策略", "TournamentBanStrategy": "竞技策略"}
        pickModes = {"": "待定", "AllRandomPickStrategy": "全随机模式", "AllTeamVotePickStrategy": "全队投票", "CounterDraftPickStrategy": "互选模式", "DraftModeSinglePickStrategy": "传统征召模式", "OneTeamVotePickStrategy": "单队投票", "QuickplayPickStrategy": "快速匹配", "SimulPickStrategy": "自选模式", "SkipPickStrategy": "跳过英雄选择", "TeamBuilderDraftPickStrategy": "征召模式", "TournamentPickStrategy": "竞技征召模式"}
        tiers = {"": "", "NONE": "没有段位", "IRON": "坚韧黑铁", "BRONZE": "英勇黄铜", "SILVER": "不屈白银", "GOLD": "荣耀黄金", "PLATINUM": "华贵铂金", "EMERALD": "流光翡翠", "DIAMOND": "璀璨钻石", "MASTER": "超凡大师", "GRANDMASTER": "傲世宗师", "CHALLENGER": "最强王者"}
        ratedTiers = {"": "", "NONE": "没有段位", "GRAY": "灰白", "GREEN": "翠绿", "BLUE": "天蓝", "PURPLE": "绛紫", "ORANGE": "耀橙"}
        tiers_all = tiers | ratedTiers
        gameTypes = {"MATCHED_GAME": "匹配对局", "CUSTOM_GAME": "自定义对局", "TUTORIAL_GAME": "新手教程"}
        team_color = {100: "蓝方", 200: "红方"}
        subteam_color = {0: "", 1: "魄罗", 2: "小兵", 3: "迅捷蟹", 4: "石甲虫", 5: "锋喙鸟", 6: "哨卫", 7: "狼", 8: "魔沼蛙"} #仅用于斗魂竞技场（Only for Arena mode）
        endOfGameResults = {"": "", "GameComplete": "游戏结束", "Abort_Unexpected": "意外终止", "Abort_TooFewPlayers": "全员提前退出", "Abort_AntiCheatExit": "检测到作弊而终止"}
        augment_rarity = {0: "白银", 1: "黄金", 2: "棱彩", 4: "黄金", 8: "棱彩", "kBronze": "青铜", "kSilver": "白银", "kGold": "黄金", "kPrismatic": "棱彩"}
        lanes = {"TOP": "上路", "JUNGLE": "打野", "MIDDLE": "中路", "BOTTOM": "下路", "NONE": ""}
        roles = {"CARRY": "C位", "DUO": "游走", "SOLO": "单人", "SUPPORT": "辅助", "NONE": ""}
        traitStyles = {0: "", 1: "青铜", 2: "白银", 3: "黄金", 4: "炫金", 5: "独特"}
        rarities = {"Default": "经典", "NoRarity": "其它", "Epic": "史诗", "Legendary": "传说", "Mythic": "神话", "Rare": "稀有", "Ultimate": "终极", "Exalted": "圣者至尊", "Transcendant": "超凡"}
        win_dict = {True: "胜利", False: "失败"}
    elif arg_locale == "en_US":
        categories = {"Custom": "Custom", "PvP": "PvP", "VersusAi": "VersusAi"}
        gameSelectCategories = {"": "", "CreateCustom": "Create Custom", "JoinCustom": "Join Custom", "kPvP": "PvP", "kTraining": "Training", "kVersusAI": "Co-op vs. AI"}
        gameSelectModeGroups = {"": "", "kARAM": "ARAM", "kAlternativeLeagueGameModes": "Alternate League Modes", "kSummonersRift": "Summoner's Rift", "kTeamfightTactics": "Teamfight Tactics"}
        queueAvailability_dict = {"Available": "√", "PlatformDisabled": ""}
        queueTypes = {"RANKED_SOLO_5x5": "Ranked Solo/Duo", "RANKED_FLEX_SR": "Ranked Flex", "RANKED_TFT": "Ranked TFT", "RANKED_TFT_PAIRS": "2V0", "RANKED_TFT_DOUBLE_UP": "Double Up", "RANKED_TFT_TURBO": "Hyper Roll", "CHERRY": "Arena"}
        banModes = {"": "", "SkipBanStrategy": "SkipBanStrategy", "StandardBanStrategy": "StandardBanStrategy", "TournamentBanStrategy": "TournamentBanStrategy"}
        pickModes = {"": "", "AllRandomPickStrategy": "AllRandomPickStrategy", "AllTeamVotePickStrategy": "AllTeamVotePickStrategy", "CounterDraftPickStrategy": "CounterDraftPickStrategy", "DraftModeSinglePickStrategy": "DraftModeSinglePickStrategy", "OneTeamVotePickStrategy": "OneTeamVotePickStrategy", "QuickplayPickStrategy": "QuickplayPickStrategy", "SimulPickStrategy": "SimulPickStrategy", "SkipPickStrategy": "SkipPickStrategy", "TeamBuilderDraftPickStrategy": "TeamBuilderDraftPickStrategy", "TournamentPickStrategy": "TournamentPickStrategy"}
        tiers = {"": "", "NONE": "NONE", "IRON": "IRON", "BRONZE": "BRONZE", "SILVER": "SILVER", "GOLD": "GOLD", "PLATINUM": "PLATINUM", "EMERALD": "EMERALD", "DIAMOND": "DIAMOND", "MASTER": "MASTER", "GRANDMASTER": "GRANDMASTER", "CHALLENGER": "CHALLENGER"}
        ratedTiers = {"": "", "NONE": "NONE", "GRAY": "GRAY", "GREEN": "GREEN", "BLUE": "BLUE", "PURPLE": "PURPLE", "ORANGE": "ORANGE"}
        tiers_all = tiers | ratedTiers
        gameTypes = {"MATCHED_GAME": "MATCHED_GAME", "CUSTOM_GAME": "CUSTOM_GAME", "TUTORIAL_GAME": "TUTORIAL_GAME"}
        team_color = {100: "BLUE", 200: "RED"}
        subteam_color = {0: "", 1: "PORO", 2: "MINION", 3: "SCUTTLE", 4: "KRUG", 5: "RAPTOR", 6: "SENTINEL", 7: "WOLF", 8: "GROMP"} #仅用于斗魂竞技场（Only for Arena mode）
        endOfGameResults = {"": "", "GameComplete": "GameComplete", "Abort_Unexpected": "Abort_Unexpected", "Abort_TooFewPlayers": "Abort_TooFewPlayers", "Abort_AntiCheatExit": "Abort_AntiCheatExit"}
        augment_rarity = {0: "Silver", 1: "Gold", 2: "Prismatic", 4: "Gold", 8: "Prismatic", "kBronze": "Bronze", "kSilver": "Silver", "kGold": "Gold", "kPrismatic": "Prismatic"}
        lanes = {"TOP": "TOP", "JUNGLE": "JUNGLE", "MIDDLE": "MIDDLE", "BOTTOM": "BOTTOM", "NONE": ""}
        roles = {"CARRY": "CARRY", "DUO": "DUO", "SOLO": "SOLO", "SUPPORT": "SUPPORT", "NONE": ""}
        traitStyles = {0: "", 1: "Bronze", 2: "Silver", 3: "Gold", 4: "Golden", 5: "Unique"}
        rarities = {"Default": "Default", "NoRarity": "Other", "Epic": "Epic", "Legendary": "Legendary", "Mythic": "Mythic", "Rare": "Rare", "Ultimate": "Ultimate", "Exalted": "Exalted", "Transcendant": "Transcendant"}
        win_dict = {True: "VICTORY", False: "DEFEAT"}
    #定义常量（Define constant values）
    global wd, bot_puuid, queueId_options, LoLGame_stat_header, LoLGame_stat_header_keys, TFTGame_stat_header, TFTGame_stat_header_keys, social_leaderboard_header, social_leaderboard_header_keys, created_processes
    wd = os.getcwd()
    bot_puuid = "00000000-0000-0000-0000-000000000000"
    queueId_options = {"#1": {"description": "所有玩家对战（All PvP games）", "expression": '[queue["id"] for queue in queues_initial if queue["category"] == "PvP"]'}, "#2": {"description": "所有英雄联盟玩家对战（All LoL PvP games）", "expression": '[queue["id"] for queue in queues_initial if queue["category"] == "PvP" and queue["mapId"] != 22]'}, "#3": {"description": "所有云顶之弈玩家对战（All TFT PvP games）", "expression": '[queue["id"] for queue in queues_initial if queue["category"] == "PvP" and queue["mapId"] == 22]'}, "#4": {"description": "所有召唤师峡谷排位队列（All Summoner's Rift ranked queues）", "expression": '[queue["id"] for queue in queues_initial if queue["isRanked"] and queue["mapId"] == 11]'}, "#5": {"description": "所有云顶之弈排位队列（All TFT ranked queues）", "expression": '[queue["id"] for queue in queues_initial if queue["isRanked"] and queue["mapId"] == 22]'}}
    LoLGame_stat_header = {"gameIndex": "游戏序号", "endOfGameResult": "对局终止情况", "gameCreation": "对局创建时间戳", "gameCreationDate": "对局创建日期", "gameDuration": "持续时长", "gameId": "对局序号", "gameMode": "游戏模式", "gameType": "游戏类型", "gameVersion": "对局版本", "mapId": "地图序号", "queueId": "队列序号", "gameModeName": "游戏模式名称", "accountId": "帐户序号", "currentAccountId": "当前帐户序号", "currentPlatformId": "当前服务器代码", "gameName": "玩家昵称", "matchHistoryUri": "对局记录网址", "platformId": "服务器代码", "profileIcon": "召唤师图标序号", "puuid": "玩家通用唯一识别码", "summonerId": "召唤师序号", "summonerName": "召唤师名称", "tagLine": "昵称编号", "profileIcon_title": "召唤师图标名称", "profileIcon_imagePath": "召唤师图标路径", "championId": "英雄序号", "highestAchievedSeasonTier": "最高段位", "participantId": "玩家序号", "spell1Id": "召唤师技能1序号", "spell2Id": "召唤师技能2序号", "teamId": "阵营", "champion_name": "英雄", "champion_alias": "代号", "champion_squarePortraitPath": "方块头像路径", "spell1_name": "召唤师技能1", "spell2_name": "召唤师技能2", "spell1_iconPath": "召唤师技能1图标", "spell2_iconPath": "召唤师技能2图标", "assists": "助攻", "causedEarlySurrender": "发起提前投降", "champLevel": "英雄等级", "combatPlayerScore": "战斗得分", "damageDealtToObjectives": "对战略点的总伤害", "damageDealtToTurrets": "对防御塔的总伤害", "damageSelfMitigated": "自我缓和的伤害", "deaths": "死亡", "doubleKills": "双杀", "earlySurrenderAccomplice": "同意提前投降", "firstBloodAssist": "协助获得第一滴血", "firstBloodKill": "第一滴血", "firstInhibitorAssist": "协助摧毁第一座召唤水晶", "firstInhibitorKill": "摧毁第一座召唤水晶", "firstTowerAssist": "协助摧毁第一座塔", "firstTowerKill": "摧毁第一座塔", "gameEndedInEarlySurrender": "提前投降导致比赛结束", "gameEndedInSurrender": "投降导致比赛结束", "goldEarned": "金币获取", "goldSpent": "金币使用", "inhibitorKills": "摧毁召唤水晶", "item0": "装备1序号", "item1": "装备2序号", "item2": "装备3序号", "item3": "装备4序号", "item4": "装备5序号", "item5": "装备6序号", "item6": "饰品序号", "killingSprees": "大杀特杀", "kills": "击杀", "largestCriticalStrike": "最大暴击伤害", "largestKillingSpree": "最高连杀", "largestMultiKill": "最高多杀", "longestTimeSpentLiving": "最长生存时间", "magicDamageDealt": "造成的魔法伤害", "magicDamageDealtToChampions": "对英雄的魔法伤害", "magicalDamageTaken": "承受的魔法伤害", "neutralMinionsKilled": "击杀野怪", "neutralMinionsKilledEnemyJungle": "击杀敌方野区野怪", "neutralMinionsKilledTeamJungle": "击杀我方野区野怪", "objectivePlayerScore": "战略点玩家得分", "pentaKills": "五杀", "perk0": "符文1序号", "perk0Var1": "符文1：参数1", "perk0Var2": "符文1：参数2", "perk0Var3": "符文1：参数3", "perk1": "符文2序号", "perk1Var1": "符文2：参数1", "perk1Var2": "符文2：参数2", "perk1Var3": "符文2：参数3", "perk2": "符文3序号", "perk2Var1": "符文3：参数1", "perk2Var2": "符文3：参数2", "perk2Var3": "符文3：参数3", "perk3": "符文4序号", "perk3Var1": "符文4：参数1", "perk3Var2": "符文4：参数2", "perk3Var3": "符文4：参数3", "perk4": "符文5序号", "perk4Var1": "符文5：参数1", "perk4Var2": "符文5：参数2", "perk4Var3": "符文5：参数3", "perk5": "符文6序号", "perk5Var1": "符文6：参数1", "perk5Var2": "符文6：参数2", "perk5Var3": "符文6：参数3", "perkPrimaryStyle": "主系序号", "perkSubStyle": "副系序号", "physicalDamageDealt": "造成的物理伤害", "physicalDamageDealtToChampions": "对英雄的物理伤害", "physicalDamageTaken": "承受的物理伤害", "playerAugment1": "强化符文1", "playerAugment2": "强化符文2", "playerAugment3": "强化符文3", "playerAugment4": "强化符文4", "playerAugment5": "强化符文5", "playerAugment6": "强化符文6", "playerScore0": "玩家得分1", "playerScore1": "玩家得分2", "playerScore2": "玩家得分3", "playerScore3": "玩家得分4", "playerScore4": "玩家得分5", "playerScore5": "玩家得分6", "playerScore6": "玩家得分7", "playerScore7": "玩家得分8", "playerScore8": "玩家得分9", "playerScore9": "玩家得分10", "playerSubteamId": "子阵营序号", "quadraKills": "四杀", "sightWardsBoughtInGame": "购买洞察之石", "subteamPlacement": "队伍排名", "teamEarlySurrendered": "队伍提前投降", "timeCCingOthers": "控制得分", "totalDamageDealt": "造成的伤害总和", "totalDamageDealtToChampions": "对英雄的伤害总和", "totalDamageTaken": "承受伤害", "totalHeal": "输出治疗效果", "totalMinionsKilled": "击杀小兵", "totalPlayerScore": "玩家总得分", "totalScoreRank": "总得分排名", "totalTimeCrowdControlDealt": "控制时间", "totalUnitsHealed": "治疗单位数", "tripleKills": "三杀", "trueDamageDealt": "造成真实伤害", "trueDamageDealtToChampions": "对英雄的真实伤害", "trueDamageTaken": "承受的真实伤害", "turretKills": "摧毁防御塔", "unrealKills": "六杀及以上", "visionScore": "视野得分", "visionWardsBoughtInGame": "购买控制守卫", "wardsKilled": "摧毁守卫", "wardsPlaced": "放置守卫", "win": "胜利", "item0_name": "装备1", "item1_name": "装备2", "item2_name": "装备3", "item3_name": "装备4", "item4_name": "装备5", "item5_name": "装备6", "item6_name": "饰品", "item0_iconPath": "装备1图标路径", "item1_iconPath": "装备2图标路径", "item2_iconPath": "装备3图标路径", "item3_iconPath": "装备4图标路径", "item4_iconPath": "装备5图标路径", "item5_iconPath": "装备6图标路径", "item6_iconPath": "饰品图标路径", "perk0EndOfGameStatDescs": "符文1游戏结算数据", "perk1EndOfGameStatDescs": "符文2游戏结算数据", "perk2EndOfGameStatDescs": "符文3游戏结算数据", "perk3EndOfGameStatDescs": "符文4游戏结算数据", "perk4EndOfGameStatDescs": "符文5游戏结算数据", "perk5EndOfGameStatDescs": "符文6游戏结算数据", "perk0_name": "符文1名称", "perk1_name": "符文2名称", "perk2_name": "符文3名称", "perk3_name": "符文4名称", "perk4_name": "符文5名称", "perk5_name": "符文6名称", "perk0_iconPath": "符文1图标路径", "perk1_iconPath": "符文2图标路径", "perk2_iconPath": "符文3图标路径", "perk3_iconPath": "符文4图标路径", "perk4_iconPath": "符文5图标路径", "perk5_iconPath": "符文6图标路径", "perkPrimaryStyle_name": "主系名称", "perkPrimaryStyle_iconPath": "主系图标路径", "perkSubStyle_name": "副系名称", "perkSubStyle_iconPath": "副系图标路径", "playerAugment1_nameTRA": "强化符文1名称", "playerAugment2_nameTRA": "强化符文2名称", "playerAugment3_nameTRA": "强化符文3名称", "playerAugment4_nameTRA": "强化符文4名称", "playerAugment5_nameTRA": "强化符文5名称", "playerAugment6_nameTRA": "强化符文6名称", "playerAugment1_augmentIconPath": "强化符文1图标路径", "playerAugment2_augmentIconPath": "强化符文2图标路径", "playerAugment3_augmentIconPath": "强化符文3图标路径", "playerAugment4_augmentIconPath": "强化符文4图标路径", "playerAugment5_augmentIconPath": "强化符文5图标路径", "playerAugment6_augmentIconPath": "强化符文6图标路径", "playerAugment1_rarity": "强化符文1等级", "playerAugment2_rarity": "强化符文2等级", "playerAugment3_rarity": "强化符文3等级", "playerAugment4_rarity": "强化符文4等级", "playerAugment5_rarity": "强化符文5等级", "playerAugment6_rarity": "强化符文6等级", "K/D/A": "击杀/死亡/助攻", "KDA": "战损比", "CS": "补刀", "GPM": "分均经济", "GUE": "金币利用率", "CSPM": "分均补刀", "D/G": "伤害转化率", "win/lose": "胜负", "lane": "分路", "role": "角色定位", "assists_percent": "助攻次数占比", "combatPlayerScore_percent": "战斗得分占比", "damageDealtToObjectives_percent": "对战略点的总伤害占比", "damageDealtToTurrets_percent": "对防御塔的总伤害占比", "damageSelfMitigated_percent": "自我缓和的伤害占比", "deaths_percent": "死亡次数占比", "doubleKills_percent": "双杀次数占比", "goldEarned_percent": "金币获取占比", "goldSpent_percent": "金币使用占比", "inhibitorKills_percent": "摧毁召唤水晶数量占比", "killingSprees_percent": "大杀特杀次数占比", "kills_percent": "击杀数量占比", "largestCriticalStrike_percent": "最大暴击伤害占比", "largestKillingSpree_percent": "最高连杀占比", "largestMultiKill_percent": "最高多杀占比", "longestTimeSpentLiving_percent": "最长生存时间占比", "magicDamageDealt_percent": "造成的魔法伤害占比", "magicDamageDealtToChampions_percent": "对英雄的魔法伤害占比", "magicalDamageTaken_percent": "承受的魔法伤害占比", "neutralMinionsKilled_percent": "击杀野怪数量占比", "neutralMinionsKilledEnemyJungle_percent": "击杀敌方野区野怪数量占比", "neutralMinionsKilledTeamJungle_percent": "击杀我方野区野怪数量占比", "objectivePlayerScore_percent": "战略点玩家得分占比", "pentaKills_percent": "五杀次数占比", "physicalDamageDealt_percent": "造成的物理伤害占比", "physicalDamageDealtToChampions_percent": "对英雄的物理伤害占比", "physicalDamageTaken_percent": "承受的物理伤害占比", "playerScore0_percent": "玩家得分1占比", "playerScore1_percent": "玩家得分2占比", "playerScore2_percent": "玩家得分3占比", "playerScore3_percent": "玩家得分4占比", "playerScore4_percent": "玩家得分5占比", "playerScore5_percent": "玩家得分6占比", "playerScore6_percent": "玩家得分7占比", "playerScore7_percent": "玩家得分8占比", "playerScore8_percent": "玩家得分9占比", "playerScore9_percent": "玩家得分10占比", "quadraKills_percent": "四杀次数占比", "sightWardsBoughtInGame_percent": "购买洞察之石数量占比", "timeCCingOthers_percent": "控制得分占比", "totalDamageDealt_percent": "造成的伤害总和占比", "totalDamageDealtToChampions_percent": "对英雄的伤害总和占比", "totalDamageTaken_percent": "承受伤害占比", "totalHeal_percent": "输出治疗效果占比", "totalMinionsKilled_percent": "击杀小兵数量占比", "totalPlayerScore_percent": "玩家总得分占比", "totalTimeCrowdControlDealt_percent": "控制时间占比", "totalUnitsHealed_percent": "治疗单位数占比", "tripleKills_percent": "三杀次数占比", "trueDamageDealt_percent": "造成真实伤害占比", "trueDamageDealtToChampions_percent": "对英雄的真实伤害占比", "trueDamageTaken_percent": "承受的真实伤害占比", "turretKills_percent": "摧毁防御塔数量占比", "unrealKills_percent": "六杀及以上连杀次数占比", "visionScore_percent": "视野得分占比", "visionWardsBoughtInGame_percent": "购买控制守卫数量占比", "wardsKilled_percent": "摧毁守卫数量占比", "wardsPlaced_percent": "放置守卫数量占比", "KP_percent": "参团率", "CS_percent": "补刀数占比", "assists_order": "助攻次数位次", "champLevel_order": "英雄等级位次", "combatPlayerScore_order": "战斗得分位次", "damageDealtToObjectives_order": "对战略点的总伤害位次", "damageDealtToTurrets_order": "对防御塔的总伤害位次", "damageSelfMitigated_order": "自我缓和的伤害位次", "deaths_order": "死亡次数位次", "doubleKills_order": "双杀次数位次", "goldEarned_order": "金币获取位次", "goldSpent_order": "金币使用位次", "inhibitorKills_order": "摧毁召唤水晶数量位次", "killingSprees_order": "大杀特杀次数位次", "kills_order": "击杀数量位次", "largestCriticalStrike_order": "最大暴击伤害位次", "largestKillingSpree_order": "最高连杀位次", "largestMultiKill_order": "最高多杀位次", "longestTimeSpentLiving_order": "最长生存时间位次", "magicDamageDealt_order": "造成的魔法伤害位次", "magicDamageDealtToChampions_order": "对英雄的魔法伤害位次", "magicalDamageTaken_order": "承受的魔法伤害位次", "neutralMinionsKilled_order": "击杀野怪数量位次", "neutralMinionsKilledEnemyJungle_order": "击杀敌方野区野怪数量位次", "neutralMinionsKilledTeamJungle_order": "击杀我方野区野怪数量位次", "objectivePlayerScore_order": "战略点玩家得分位次", "pentaKills_order": "五杀次数位次", "physicalDamageDealt_order": "造成的物理伤害位次", "physicalDamageDealtToChampions_order": "对英雄的物理伤害位次", "physicalDamageTaken_order": "承受的物理伤害位次", "playerScore0_order": "玩家得分1位次", "playerScore1_order": "玩家得分2位次", "playerScore2_order": "玩家得分3位次", "playerScore3_order": "玩家得分4位次", "playerScore4_order": "玩家得分5位次", "playerScore5_order": "玩家得分6位次", "playerScore6_order": "玩家得分7位次", "playerScore7_order": "玩家得分8位次", "playerScore8_order": "玩家得分9位次", "playerScore9_order": "玩家得分10位次", "quadraKills_order": "四杀次数位次", "sightWardsBoughtInGame_order": "购买洞察之石数量位次", "timeCCingOthers_order": "控制得分位次", "totalDamageDealt_order": "造成的伤害总和位次", "totalDamageDealtToChampions_order": "对英雄的伤害总和位次", "totalDamageTaken_order": "承受伤害位次", "totalHeal_order": "输出治疗效果位次", "totalMinionsKilled_order": "击杀小兵数量位次", "totalPlayerScore_order": "玩家总得分位次", "totalTimeCrowdControlDealt_order": "控制时间位次", "totalUnitsHealed_order": "治疗单位数位次", "tripleKills_order": "三杀次数位次", "trueDamageDealt_order": "造成真实伤害位次", "trueDamageDealtToChampions_order": "对英雄的真实伤害位次", "trueDamageTaken_order": "承受的真实伤害位次", "turretKills_order": "摧毁防御塔数量位次", "unrealKills_order": "六杀及以上连杀次数位次", "visionScore_order": "视野得分位次", "visionWardsBoughtInGame_order": "购买控制守卫数量位次", "wardsKilled_order": "摧毁守卫数量位次", "wardsPlaced_order": "放置守卫数量位次", "KDA_order": "战损比位次", "KP_order": "参团率位次", "CS_order": "补刀数位次", "D/G_order": "伤害转化率位次", "GUE_order": "金币利用率位次"}
    LoLGame_stat_header_keys = list(LoLGame_stat_header.keys())
    TFTGame_stat_header = {"gameIndex": "游戏序号", "endOfGameResult": "对局终止情况", "gameCreation": "对局创建时间", "game_datetime": "对局结算时间", "game_id": "对局序号", "game_length": "持续时长", "game_version": "对局版本", "queue_id": "队列序号", "tft_game_type": "游戏类型", "tft_set_core_name": "数据版本名称", "tft_set_number": "赛季", "participantId": "玩家序号", "augment1 apiName": "强化符文1接口名称", "augment2 apiName": "强化符文2接口名称", "augment3 apiName": "强化符文3接口名称", "augment1 name": "强化符文1名称", "augment2 name": "强化符文2名称", "augment3 name": "强化符文3名称", "augment1 icon": "强化符文1图标", "augment2 icon": "强化符文2图标", "augment3 icon": "强化符文3图标", "companion content_ID": "小小英雄商品编号", "companion item_ID": "小小英雄序号", "companion skin_ID": "小小英雄皮肤序号", "companion species": "小小英雄物种", "companion name": "小小英雄名称", "companion level": "小小英雄星级", "companion rarity": "小小英雄稀有度", "gold_left": "剩余金币", "last_round": "存活回合", "level": "等级", "placement": "名次", "players_eliminated": "淘汰玩家数", "puuid": "玩家通用唯一识别码", "time_eliminated": "存活时长", "total_damage_to_players": "造成玩家伤害", "summonerId": "召唤师序号", "summonerName": "召唤师名称", "gameName": "玩家昵称", "tagLine": "昵称编号", "trait0 name": "羁绊1", "trait0 num_units": "羁绊1单位数", "trait0 style": "羁绊1羁绊框颜色", "trait0 tier_current": "羁绊1当前等级", "trait0 tier_total": "羁绊1最高等级", "trait0 display_name": "羁绊1显示名", "trait0 icon_path": "羁绊1图标路径", "trait1 name": "羁绊2", "trait1 num_units": "羁绊2单位数", "trait1 style": "羁绊2羁绊框颜色", "trait1 tier_current": "羁绊2当前等级", "trait1 tier_total": "羁绊2最高等级", "trait1 display_name": "羁绊2显示名", "trait1 icon_path": "羁绊2图标路径", "trait2 name": "羁绊3", "trait2 num_units": "羁绊3单位数", "trait2 style": "羁绊3羁绊框颜色", "trait2 tier_current": "羁绊3当前等级", "trait2 tier_total": "羁绊3最高等级", "trait2 display_name": "羁绊3显示名", "trait2 icon_path": "羁绊3图标路径", "trait3 name": "羁绊4", "trait3 num_units": "羁绊4单位数", "trait3 style": "羁绊4羁绊框颜色", "trait3 tier_current": "羁绊4当前等级", "trait3 tier_total": "羁绊4最高等级", "trait3 display_name": "羁绊4显示名", "trait3 icon_path": "羁绊4图标路径", "trait4 name": "羁绊5", "trait4 num_units": "羁绊5单位数", "trait4 style": "羁绊5羁绊框颜色", "trait4 tier_current": "羁绊5当前等级", "trait4 tier_total": "羁绊5最高等级", "trait4 display_name": "羁绊5显示名", "trait4 icon_path": "羁绊5图标路径", "trait5 name": "羁绊6", "trait5 num_units": "羁绊6单位数", "trait5 style": "羁绊6羁绊框颜色", "trait5 tier_current": "羁绊6当前等级", "trait5 tier_total": "羁绊6最高等级", "trait5 display_name": "羁绊6显示名", "trait5 icon_path": "羁绊6图标路径", "trait6 name": "羁绊7", "trait6 num_units": "羁绊7单位数", "trait6 style": "羁绊7羁绊框颜色", "trait6 tier_current": "羁绊7当前等级", "trait6 tier_total": "羁绊7最高等级", "trait6 display_name": "羁绊7显示名", "trait6 icon_path": "羁绊7图标路径", "trait7 name": "羁绊8", "trait7 num_units": "羁绊8单位数", "trait7 style": "羁绊8羁绊框颜色", "trait7 tier_current": "羁绊8当前等级", "trait7 tier_total": "羁绊8最高等级", "trait7 display_name": "羁绊8显示名", "trait7 icon_path": "羁绊8图标路径", "trait8 name": "羁绊9", "trait8 num_units": "羁绊9单位数", "trait8 style": "羁绊9羁绊框颜色", "trait8 tier_current": "羁绊9当前等级", "trait8 tier_total": "羁绊9最高等级", "trait8 display_name": "羁绊9显示名", "trait8 icon_path": "羁绊9图标路径", "trait9 name": "羁绊10", "trait9 num_units": "羁绊10单位数", "trait9 style": "羁绊10羁绊框颜色", "trait9 tier_current": "羁绊10当前等级", "trait9 tier_total": "羁绊10最高等级", "trait9 display_name": "羁绊10显示名", "trait9 icon_path": "羁绊10图标路径", "trait10 name": "羁绊11", "trait10 num_units": "羁绊11单位数", "trait10 style": "羁绊11羁绊框颜色", "trait10 tier_current": "羁绊11当前等级", "trait10 tier_total": "羁绊11最高等级", "trait10 display_name": "羁绊11显示名", "trait10 icon_path": "羁绊11图标路径", "trait11 name": "羁绊12", "trait11 num_units": "羁绊12单位数", "trait11 style": "羁绊12羁绊框颜色", "trait11 tier_current": "羁绊12当前等级", "trait11 tier_total": "羁绊12最高等级", "trait11 display_name": "羁绊12显示名", "trait11 icon_path": "羁绊12图标路径", "trait12 name": "羁绊13", "trait12 num_units": "羁绊13单位数", "trait12 style": "羁绊13羁绊框颜色", "trait12 tier_current": "羁绊13当前等级", "trait12 tier_total": "羁绊13最高等级", "trait12 display_name": "羁绊13显示名", "trait12 icon_path": "羁绊13图标路径", "unit0 character_id": "英雄1：角色编号", "unit0 rarity": "英雄1：卡费", "unit0 tier": "英雄1：星级", "unit0 display_name": "英雄1：显示名", "unit0 squareIconPath": "英雄1：方块图标路径", "unit1 character_id": "英雄2：角色编号", "unit1 rarity": "英雄2：卡费", "unit1 tier": "英雄2：星级", "unit1 display_name": "英雄2：显示名", "unit1 squareIconPath": "英雄2：方块图标路径", "unit2 character_id": "英雄3：角色编号", "unit2 rarity": "英雄3：卡费", "unit2 tier": "英雄3：星级", "unit2 display_name": "英雄3：显示名", "unit2 squareIconPath": "英雄3：方块图标路径", "unit3 character_id": "英雄4：角色编号", "unit3 rarity": "英雄4：卡费", "unit3 tier": "英雄4：星级", "unit3 display_name": "英雄4：显示名", "unit3 squareIconPath": "英雄4：方块图标路径", "unit4 character_id": "英雄5：角色编号", "unit4 rarity": "英雄5：卡费", "unit4 tier": "英雄5：星级", "unit4 display_name": "英雄5：显示名", "unit4 squareIconPath": "英雄5：方块图标路径", "unit5 character_id": "英雄6：角色编号", "unit5 rarity": "英雄6：卡费", "unit5 tier": "英雄6：星级", "unit5 display_name": "英雄6：显示名", "unit5 squareIconPath": "英雄6：方块图标路径", "unit6 character_id": "英雄7：角色编号", "unit6 rarity": "英雄7：卡费", "unit6 tier": "英雄7：星级", "unit6 display_name": "英雄7：显示名", "unit6 squareIconPath": "英雄7：方块图标路径", "unit7 character_id": "英雄8：角色编号", "unit7 rarity": "英雄8：卡费", "unit7 tier": "英雄8：星级", "unit7 display_name": "英雄8：显示名", "unit7 squareIconPath": "英雄8：方块图标路径", "unit8 character_id": "英雄9：角色编号", "unit8 rarity": "英雄9：卡费", "unit8 tier": "英雄9：星级", "unit8 display_name": "英雄9：显示名", "unit8 squareIconPath": "英雄9：方块图标路径", "unit9 character_id": "英雄10：角色编号", "unit9 rarity": "英雄10：卡费", "unit9 tier": "英雄10：星级", "unit9 display_name": "英雄10：显示名", "unit9 squareIconPath": "英雄10：方块图标路径", "unit10 character_id": "英雄11：角色编号", "unit10 rarity": "英雄11：卡费", "unit10 tier": "英雄11：星级", "unit10 display_name": "英雄11：显示名", "unit10 squareIconPath": "英雄11：方块图标路径", "unit0 item0 nameId": "英雄1：装备1序号", "unit0 item0 name": "英雄1：装备1名称", "unit0 item0 squareIconPath": "英雄1：装备1方块图像路径", "unit0 item1 nameId": "英雄1：装备2序号", "unit0 item1 name": "英雄1：装备2名称", "unit0 item1 squareIconPath": "英雄1：装备2方块图像路径", "unit0 item2 nameId": "英雄1：装备3序号", "unit0 item2 name": "英雄1：装备3名称", "unit0 item2 squareIconPath": "英雄1：装备3方块图像路径", "unit1 item0 nameId": "英雄2：装备1序号", "unit1 item0 name": "英雄2：装备1名称", "unit1 item0 squareIconPath": "英雄2：装备1方块图像路径", "unit1 item1 nameId": "英雄2：装备2序号", "unit1 item1 name": "英雄2：装备2名称", "unit1 item1 squareIconPath": "英雄2：装备2方块图像路径", "unit1 item2 nameId": "英雄2：装备3序号", "unit1 item2 name": "英雄2：装备3名称", "unit1 item2 squareIconPath": "英雄2：装备3方块图像路径", "unit2 item0 nameId": "英雄3：装备1序号", "unit2 item0 name": "英雄3：装备1名称", "unit2 item0 squareIconPath": "英雄3：装备1方块图像路径", "unit2 item1 nameId": "英雄3：装备2序号", "unit2 item1 name": "英雄3：装备2名称", "unit2 item1 squareIconPath": "英雄3：装备2方块图像路径", "unit2 item2 nameId": "英雄3：装备3序号", "unit2 item2 name": "英雄3：装备3名称", "unit2 item2 squareIconPath": "英雄3：装备3方块图像路径", "unit3 item0 nameId": "英雄4：装备1序号", "unit3 item0 name": "英雄4：装备1名称", "unit3 item0 squareIconPath": "英雄4：装备1方块图像路径", "unit3 item1 nameId": "英雄4：装备2序号", "unit3 item1 name": "英雄4：装备2名称", "unit3 item1 squareIconPath": "英雄4：装备2方块图像路径", "unit3 item2 nameId": "英雄4：装备3序号", "unit3 item2 name": "英雄4：装备3名称", "unit3 item2 squareIconPath": "英雄4：装备3方块图像路径", "unit4 item0 nameId": "英雄5：装备1序号", "unit4 item0 name": "英雄5：装备1名称", "unit4 item0 squareIconPath": "英雄5：装备1方块图像路径", "unit4 item1 nameId": "英雄5：装备2序号", "unit4 item1 name": "英雄5：装备2名称", "unit4 item1 squareIconPath": "英雄5：装备2方块图像路径", "unit4 item2 nameId": "英雄5：装备3序号", "unit4 item2 name": "英雄5：装备3名称", "unit4 item2 squareIconPath": "英雄5：装备3方块图像路径", "unit5 item0 nameId": "英雄6：装备1序号", "unit5 item0 name": "英雄6：装备1名称", "unit5 item0 squareIconPath": "英雄6：装备1方块图像路径", "unit5 item1 nameId": "英雄6：装备2序号", "unit5 item1 name": "英雄6：装备2名称", "unit5 item1 squareIconPath": "英雄6：装备2方块图像路径", "unit5 item2 nameId": "英雄6：装备3序号", "unit5 item2 name": "英雄6：装备3名称", "unit5 item2 squareIconPath": "英雄6：装备3方块图像路径", "unit6 item0 nameId": "英雄7：装备1序号", "unit6 item0 name": "英雄7：装备1名称", "unit6 item0 squareIconPath": "英雄7：装备1方块图像路径", "unit6 item1 nameId": "英雄7：装备2序号", "unit6 item1 name": "英雄7：装备2名称", "unit6 item1 squareIconPath": "英雄7：装备2方块图像路径", "unit6 item2 nameId": "英雄7：装备3序号", "unit6 item2 name": "英雄7：装备3名称", "unit6 item2 squareIconPath": "英雄7：装备3方块图像路径", "unit7 item0 nameId": "英雄8：装备1序号", "unit7 item0 name": "英雄8：装备1名称", "unit7 item0 squareIconPath": "英雄8：装备1方块图像路径", "unit7 item1 nameId": "英雄8：装备2序号", "unit7 item1 name": "英雄8：装备2名称", "unit7 item1 squareIconPath": "英雄8：装备2方块图像路径", "unit7 item2 nameId": "英雄8：装备3序号", "unit7 item2 name": "英雄8：装备3名称", "unit7 item2 squareIconPath": "英雄8：装备3方块图像路径", "unit8 item0 nameId": "英雄9：装备1序号", "unit8 item0 name": "英雄9：装备1名称", "unit8 item0 squareIconPath": "英雄9：装备1方块图像路径", "unit8 item1 nameId": "英雄9：装备2序号", "unit8 item1 name": "英雄9：装备2名称", "unit8 item1 squareIconPath": "英雄9：装备2方块图像路径", "unit8 item2 nameId": "英雄9：装备3序号", "unit8 item2 name": "英雄9：装备3名称", "unit8 item2 squareIconPath": "英雄9：装备3方块图像路径", "unit9 item0 nameId": "英雄10：装备1序号", "unit9 item0 name": "英雄10：装备1名称", "unit9 item0 squareIconPath": "英雄10：装备1方块图像路径", "unit9 item1 nameId": "英雄10：装备2序号", "unit9 item1 name": "英雄10：装备2名称", "unit9 item1 squareIconPath": "英雄10：装备2方块图像路径", "unit9 item2 nameId": "英雄10：装备3序号", "unit9 item2 name": "英雄10：装备3名称", "unit9 item2 squareIconPath": "英雄10：装备3方块图像路径", "unit10 item0 nameId": "英雄11：装备1序号", "unit10 item0 name": "英雄11：装备1名称", "unit10 item0 squareIconPath": "英雄11：装备1方块图像路径", "unit10 item1 nameId": "英雄11：装备2序号", "unit10 item1 name": "英雄11：装备2名称", "unit10 item1 squareIconPath": "英雄11：装备2方块图像路径", "unit10 item2 nameId": "英雄11：装备3序号", "unit10 item2 name": "英雄11：装备3名称", "unit10 item2 squareIconPath": "英雄11：装备3方块图像路径"}
    TFTGame_stat_header_keys = list(TFTGame_stat_header.keys())
    social_leaderboard_header = {"puuid": "玩家通用唯一识别码", "displayName": "显示名", "gameName": "玩家昵称", "tagLine": "昵称编号", "division": "分级", "isProvisional": "定位中", "leaguePoints": "胜点", "losses": "负场", "miniSeriesProgress": "定位赛/晋级赛进展", "provisionalGameThreshold": "总定位场次", "provisionalGamesRemaining": "剩余定位场次", "queueType": "战区", "ratedRating": "排名分", "ratedTier": "段位", "tier": "段位", "wins": "胜场", "tier / ratedTier": "段位", "leaguePoints / ratedRating": "胜点", "timestamp": "获取时间戳", "time": "获取时间"}
    social_leaderboard_header_keys = list(social_leaderboard_header.keys())
    created_processes = [] #标记清理残留进程（Stores processes to clear at the end of the program）
    
async def get_queue_data(connection) -> pandas.DataFrame:
    queues = await (await connection.request("GET", "/lol-game-queues/v1/queues")).json()
    #以前含有"最大召唤师等级"参数（There was previously a parameter: maxLevel）
    queues_header = {"allowablePremadeSizes": "可用预组队规模", "areFreeChampionsAllowed": "允许使用周免英雄", "assetMutator": "游戏模式配置", "category": "对局类型", "championsRequiredToPlay": "需要英雄数量", "description": "游戏模式描述", "detailedDescription": "补充描述", "gameMode": "游戏模式", "gameSelectCategory": "游戏选择类别", "gameSelectModeGroup": "游戏模式分组", "gameSelectPriority": "游戏选择优先级", "hidePlayerPosition": "隐藏玩家位置", "id": "队列序号", "isCustom": "自定义对局", "isRanked": "排位赛", "isSkillTreeQueue": "技巧加成队列", "isTeamBuilderManaged": "服从阵容匹配机制", "isVisible": "客户端可见性", "lastToggledOffTime": "上次关闭时间", "lastToggledOnTime": "上次开放时间", "mapId": "地图序号", "maxDivisionForPremadeSize2": "双排最高分级限制", "maxLobbySpectatorCount": "房间最大观战者数量", "maxTierForPremadeSize2": "双排最高段位限制", "maximumParticipantListSize": "最大玩家数量", "minLevel": "最低召唤师等级", "minimumParticipantListSize": "最小玩家数量", "name": "游戏模式名称", "numPlayersPerTeam": "队伍规模", "numberOfTeamsInLobby": "房间内队伍数量", "queueAvailability": "队列可用性", "removalFromGameAllowed": "允许退出游戏", "removalFromGameDelayMinutes": "允许退出游戏时间（分钟）", "shortName": "游戏模式简称", "showPositionSelector": "呈现位置指示器", "showQuickPlaySlotSelection": "呈现快速模式偏好英雄选择界面", "spectatorEnabled": "允许观战", "type": "游戏类型", "advancedLearningQuests": "进阶教程", "allowTrades": "允许交换", "banMode": "禁用模式", "banTimerDuration": "禁用时间限制（秒）", "battleBoost": "战斗加成", "crossTeamChampionPool": "跨队伍英雄共享", "deathMatch": "团体竞赛", "doNotRemove": "禁止退出游戏", "duplicatePick": "克隆选择", "exclusivePick": "唯一选择", "gameModeOverride": "游戏类型重写来源", "typeId": "游戏类型序号", "learningQuests": "新手教程", "mainPickTimerDuration": "盲选时间限制（秒）", "maxAllowableBans": "最大禁用数量", "typeName": "英雄选择策略", "numPlayersPerTeamOverride": "队伍规模重写历史", "onboardCoopBeginner": "人机对战引导模式", "pickMode": "英雄选择模式", "postPickTimerDuration": "符文和皮肤选择时间限制（秒）", "reroll": "允许重随", "teamChampionPool": "队伍英雄共享", "isChampionPointsEnabled": "队列奖励：英雄成就点数", "isIpEnabled": "队列奖励：成就", "isXpEnabled": "队列奖励：经验点数", "partySizeIpRewards": "组队额外成就奖励"}
    queues_data = {}
    queues_header_keys = list(queues_header.keys())
    for i in range(len(queues_header)):
        key = queues_header_keys[i]
        queues_data[key] = []
    for queue in queues:
        for i in range(len(queues_header_keys)):
            key = queues_header_keys[i]
            if i <= 37:
                if i == 3: #对局类型（`category`）
                    queues_data[key].append(categories[queue[key]])
                elif i == 8:
                    queues_data[key].append(gameSelectCategories[queue[key]])
                elif i == 9:
                    queues_data[key].append(gameSelectModeGroups[queue[key]])
                elif i == 18 or i == 19: #上次关闭时间和上次开放时间（`lastToggledOffTime` and `lastToggledOnTime`）
                    t = time.localtime(queue[key] / 1000)
                    standard_time = time.strftime("%Y年%m月%d日%H:%M:%S", t)
                    queues_data[key].append(standard_time)
                elif i == 23: #双排最高段位限制（`maxTierForPremadeSize2`）
                    queues_data[key].append(tiers[queue[key]])
                elif i == 30: #队列可用性（`queueAvailability`）
                    queues_data[key].append(queueAvailability_dict[queue[key]])
                else:
                    queues_data[key].append(queue[key])
            elif i <= 59:
                if i == 40: #禁用模式（`banMode`）
                    queues_data[key].append(banModes[queue["gameTypeConfig"][key]])
                elif i == 49: #游戏类型序号（`typeId`）
                    queues_data[key].append(queue["gameTypeConfig"]["id"])
                elif i == 53: #英雄选择策略（`typeName`）
                    queues_data[key].append(queue["gameTypeConfig"]["name"])
                elif i == 56: #英雄选择模式（`pickMode`）
                    queues_data[key].append(pickModes[queue["gameTypeConfig"][key]])
                else:
                    queues_data[key].append(queue["gameTypeConfig"][key])
            else:
                queues_data[key].append(queue["queueRewards"][key])
    queues_df = pandas.DataFrame(data = queues_data).sort_values(by = "id", ascending = True, ignore_index = True)
    for column in queues_df:
        if queues_df[column].dtype == "bool":
            queues_df[column] = queues_df[column].astype(str)
            for i in range(len(queues_df)):
                queues_df.loc[i, column] = "√" if queues_df[column][i] == "True" else ""
    queues_df = pandas.concat([pandas.DataFrame([queues_header])[queues_df.columns], queues_df], ignore_index = True)
    return queues_df

async def sort_social_leaderboard(connection, puuids: list) -> pandas.DataFrame:
    if all(map(verify_uuid, puuids)):
        challenger_ladder_queueTypes = await (await connection.request("GET", "/lol-ranked/v1/challenger-ladders-enabled")).json()
        topRated_ladder_queueTypes = await (await connection.request("GET", "/lol-ranked/v1/top-rated-ladders-enabled")).json()
        ranked_queueTypes = challenger_ladder_queueTypes + topRated_ladder_queueTypes
        social_leaderboard_data = {}
        for i in range(len(social_leaderboard_header_keys)):
            key = social_leaderboard_header_keys[i]
            social_leaderboard_data[key] = []
        for queueType in ranked_queueTypes:
            social_leaderboard = await (await connection.request("GET", "/lol-ranked/v1/social-leaderboard-ranked-queue-stats-for-puuids?queueType=%s&puuids=%s" %(queueType, str(puuids).replace(" ", "").replace("'", '"')))).json()
            for participant_puuid_iter in social_leaderboard:
                participant_leaderboard = social_leaderboard[participant_puuid_iter]
                participantInfo = await get_info(connection, participant_puuid_iter)
                if participantInfo["info_got"]:
                    participantInfo_body = participantInfo["body"]
                    for i in range(len(social_leaderboard_header_keys)):
                        key = social_leaderboard_header_keys[i]
                        if i <= 3:
                            social_leaderboard_data[key].append(participantInfo_body[key])
                        elif i <= 15:
                            if i == 4: #分级（`division`）
                                social_leaderboard_data[key].append("" if participant_leaderboard[key] == "NA" else participant_leaderboard[key])
                            elif i == 11: #战区（`queueType`）
                                social_leaderboard_data[key].append(queueTypes[participant_leaderboard[key]])
                            elif i == 13: #段位（`ratedTier`）
                                social_leaderboard_data[key].append(ratedTiers[participant_leaderboard[key]])
                            elif i == 14: #段位（`tier`）
                                social_leaderboard_data[key].append(tiers[participant_leaderboard[key]])
                            else:
                                social_leaderboard_data[key].append(participant_leaderboard[key])
                        elif i == 16: #段位（`tier / ratedTier`）
                            social_leaderboard_data[key].append(ratedTiers[participant_leaderboard["ratedTier"]] if queueType in topRated_ladder_queueTypes else tiers[participant_leaderboard["tier"]])
                        elif i == 17: #胜点（`leaguePoints / ratedRating`）
                            social_leaderboard_data[key].append(participant_leaderboard["ratedRating"] if queueType in topRated_ladder_queueTypes else participant_leaderboard["leaguePoints"])
                        elif i == 18: #获取时间戳（`timestamp`）
                            social_leaderboard_data[key].append(time.time())
                        else: #获取时间（`time`）
                            social_leaderboard_data[key].append(time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime()))
                else:
                    if print_detail:
                        print(participantInfo["message"])
        social_leaderboard_df = pandas.DataFrame(data = social_leaderboard_data)
        for column in social_leaderboard_df:
            if social_leaderboard_df[column].dtype == "bool":
                social_leaderboard_df[column] = social_leaderboard_df[column].astype(str)
                for i in range(len(social_leaderboard_df)):
                    social_leaderboard_df.loc[i, column] = "√" if social_leaderboard_df[column][i] == "True" else ""
        social_leaderboard_df = pandas.concat([pandas.DataFrame([social_leaderboard_header])[social_leaderboard_df.columns], social_leaderboard_df], ignore_index = True)
    else:
        social_leaderboard_df = pandas.DataFrame(data = social_leaderboard_header, index = [0])
    return social_leaderboard_df

async def search_player_match_stats_lol(connection, puuid: str, begIndex: int = 0, endIndex: int = 20) -> pandas.DataFrame: #查询某个玩家的对局记录数据（Search for a player's match history stats）
    if verify_uuid(puuid):
        info = await get_info(connection, puuid)
        if info["info_got"]:
            #准备数据资源（Prepare data resources）
            ##英雄联盟对局记录（LoL match history）
            while True:
                LoLHistory = await (await connection.request("GET", f"/lol-match-history/v1/products/lol/{puuid}/matches?begIndex={begIndex}&endIndex={endIndex}")).json()
                count = 0
                error_occurred = False
                if "errorCode" in LoLHistory:
                    if "500 Internal Server Error" in LoLHistory["message"]:
                        if not error_occurred and print_detail:
                            print("您所在大区的对局记录服务异常。尝试重新获取数据……\nThe match history service provided on your server isn't in place. Trying to recapture the history data ...")
                            error_occurred = True
                        while "errorCode" in LoLHistory and "500 Internal Server Error" in LoLHistory["message"] and count <= 3:
                            count += 1
                            if print_detail:
                                print("正在进行第%d次尝试……\nTimes trying: No. %d ..." %(count, count))
                            LoLHistory = await (await connection.request("GET", f"/lol-match-history/v1/products/lol/{puuid}/matches?begIndex={begIndex}&endIndex={endIndex}")).json()
                    elif "body was empty" in LoLHistory["message"]:
                        print("这位召唤师从5月1日起就没有进行过任何英雄联盟对局。\nThis summoner hasn't played any LoL game yet since May 1st.")
                        break
                    elif "Error getting match list for summoner" in LoLHistory["message"]:
                        LoLHistory_url = "%s/lol-match-history/v1/products/lol/%s/matches?begIndex=%d&endIndex=%d" %(connection.address, puuid, begIndex, endIndex)
                        print("请打开以下网址，输入如下所示的用户名和密码，打开后在命令行中按回车键继续（Please open the following website, type in the username and password accordingly and press Enter to continue）：\n网址（URL）：\t\t%s\n用户名（Username）：\triot\n密码（Password）：\t%s" %(LoLHistory_url, connection.auth_key))
                    if count > 3:
                        print("英雄联盟对局记录获取失败！请等待官方修复对局记录服务！\nLoL match history capture failure! Please wait for Tencent to fix the match history service!")
                        break
                else:
                    break
            if "errorCode" in LoLHistory:
                return pandas.DataFrame(data = LoLGame_stat_header, index = [0])
            #数据整理核心部分（Data sorting - core part）
            LoLGame_stat_data = {}
            for i in range(len(LoLGame_stat_header_keys)):
                key = LoLGame_stat_header_keys[i]
                LoLGame_stat_data[key] = []
            for gameIndex in range(len(LoLHistory["games"]["games"])):
                game = LoLHistory["games"]["games"][gameIndex]
                stats = game["participants"][0]["stats"]
                LoLGame_info = await (await connection.request("GET", "/lol-match-history/v1/games/%s" %(game["gameId"]))).json() #这里默认对局记录中的对局一定有相应的对局信息，因此不需要考虑其异常（Here the principle is, a match recorded in the match history must have the corresponding information, so the program doesn't need to handle any possible exception）
                team_participants = [participant for participant in LoLGame_info["participants"] if game["gameMode"] == "CHERRY" and participant["stats"]["playerSubteamId"] == stats["playerSubteamId"] or game["gameMode"] != "CHERRY" and participant["teamId"] == game["participants"][0]["teamId"]] #存储对局信息中同一队伍的玩家。斗魂竞技场对局应该使用子阵营（Store the participants of the same team from the game information. Subteam should be used to evaluate a player）
                for i in range(len(LoLGame_stat_header_keys)):
                    key = LoLGame_stat_header_keys[i]
                    if i == 0: #游戏序号（`gameIndex`）
                        LoLGame_stat_data[key].append(begIndex + 1 + gameIndex)
                    elif i <= 11:
                        if i == 1: #对局终止情况（`endOfGameResult`）
                            LoLGame_stat_data[key].append(endOfGameResults[game[key]])
                        elif i == 3: #对局创建日期（`gameCreationDate`）
                            LoLGame_stat_data[key].append(game["gameCreationDate"][:10] + " " + game["gameCreationDate"][11:23])
                        elif i == 4: #持续时长（`gameDuration`）
                            LoLGame_stat_data[key].append(str(game["gameDuration"] // 60) + ":" + "%02d" %(game["gameDuration"] % 60))
                        elif i == 7: #游戏类型（`gameType`）
                            LoLGame_stat_data[key].append(gameTypes[game[key]])
                        elif i == 11: #游戏模式名称（`gameModeName`）
                            LoLGame_stat_data[key].append("自定义" if game["queueId"] == 0 else gamemodes[game["queueId"]]["name"])
                        else:
                            LoLGame_stat_data[key].append(game[key])
                    elif i <= 24:
                        if i >= 23: #召唤师图标相关键（Summoner icon-related keys）
                            profileIconId = game["participantIdentities"][0]["player"]["profileIcon"]
                            LoLGame_stat_data[key].append(summonerIcons[profileIconId].get(key.split("_")[1], "") if profileIconId in summonerIcons else "")
                        else:
                            LoLGame_stat_data[key].append(game["participantIdentities"][0]["player"][key])
                    elif i <= 37:
                        if i == 26: #最高段位（`highestAchievedSeasonTier`）
                            LoLGame_stat_data[key].append(tiers[game["participants"][0][key]])
                        elif i == 30: #阵营（`teamId`）
                            LoLGame_stat_data[key].append(team_color[game["participants"][0][key]])
                        elif i >= 31 and i <= 33: #英雄相关键（Champion-related keys）
                            LoLGame_stat_data[key].append(LoLChampions[game["participants"][0]["championId"]][key.split("_")[1]] if game["participants"][0]["championId"] in LoLChampions else "")
                        elif i >= 34: #召唤师技能相关键（Summoner spell-related keys）
                            spellId = game["participants"][0][key.split("_")[0] + "Id"]
                            LoLGame_stat_data[key].append(spells[spellId][key.split("_")[1]] if spellId in spells else "")
                        else:
                            LoLGame_stat_data[key].append(game["participants"][0][key])
                    elif i <= 212:
                        if i == 125: #子阵营序号（`playerSubteamId`）
                            LoLGame_stat_data[key].append(subteam_color[stats[key]])
                        elif i >= 151 and i <= 164: #英雄联盟装备相关键（LoLItem-related keys）
                            itemId = stats[key.split("_")[0]]
                            LoLGame_stat_data[key].append(LoLItems[itemId][key.split("_")[1]] if itemId in LoLItems else "")
                        elif i >= 165 and i <= 182: #符文相关键（Perk-related keys）
                            if i <= 170:
                                perkId = stats[key[:5]]
                                if perkId == 0:
                                    LoLGame_stat_data[key].append("")
                                elif perkId in perks:
                                    perk_EndOfGameStatDescs = "".join(list(map(lambda x: x + "。", perks[perkId]["endOfGameStatDescs"])))
                                    perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar1@", str(stats[key[:5] + "Var1"]))
                                    perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar2@", str(stats[key[:5] + "Var2"]))
                                    perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar3@", str(stats[key[:5] + "Var3"]))
                                    LoLGame_stat_data[key].append(perk_EndOfGameStatDescs)
                                else:
                                    LoLGame_stat_data[key].append("")
                            else:
                                perkId = stats[key.split("_")[0]]
                                LoLGame_stat_data[key].append(perks[perkId][key.split("_")[1]] if perkId in perks else "")
                        elif i >= 183 and i <= 186: #符文系相关键（Perkstyles-related keys）
                            perkstyleId = stats[key.split("_")[0]]
                            LoLGame_stat_data[key].append(perkstyles[perkstyleId][key.split("_")[1]] if perkstyleId in perkstyles else "")
                        elif i >= 187 and i <= 204: #强化符文相关键（Augment-related keys）
                            CherryAugmentId = stats[key.split("_")[0]]
                            if CherryAugmentId == 0:
                                LoLGame_stat_data[key].append("")
                            elif CherryAugmentId in CherryAugments:
                                if i <= 192: #强化符文名称（`nameTRA`）
                                    LoLGame_stat_data[key].append(CherryAugments[CherryAugmentId][key.split("_")[1]])
                                elif i <= 198: #强化符文图标路径（`augmentIconPath`）
                                    LoLGame_stat_data[key].append(CherryAugments[CherryAugmentId]["augmentSmallIconPath"].replace("_small.png", "_large.png"))
                                else: #强化符文等级（`rarity`）
                                    LoLGame_stat_data[key].append(augment_rarity[CherryAugments[CherryAugmentId][key.split("_")[1]]])
                            else:
                                LoLGame_stat_data[key].append("")
                        elif i == 205: #击杀/死亡/助攻（`K/D/A`）
                            LoLGame_stat_data[key].append("/".join([str(stats["kills"]), str(stats["deaths"]), str(stats["assists"])]))
                        elif i == 206: #战损比（`KDA`）
                            LoLGame_stat_data[key].append((stats["kills"] + stats["assists"]) / max(1, stats["deaths"]))
                        elif i == 207: #补刀（`CS`）
                            LoLGame_stat_data[key].append(stats["neutralMinionsKilled"] + stats["totalMinionsKilled"])
                        elif i == 208: #分均经济（`GPM`）
                            LoLGame_stat_data[key].append(0 if game["gameDuration"] == 0 else stats["goldEarned"] * 60 / game["gameDuration"])
                        elif i == 209: #金币利用率（`GUE` - Gold Utilization Efficiency）
                            LoLGame_stat_data[key].append(0 if stats["goldEarned"] == 0 else stats["goldSpent"] / stats["goldEarned"])
                        elif i == 210: #分均补刀（`CSPM`）
                            LoLGame_stat_data[key].append(0 if game["gameDuration"] == 0 else (stats["neutralMinionsKilled"] + stats["totalMinionsKilled"]) * 60 / game["gameDuration"])
                        elif i == 211: #伤害转化率（`D/G`）
                            LoLGame_stat_data[key].append(0 if stats["goldEarned"] == 0 else stats["totalDamageDealtToChampions"] / stats["goldEarned"])
                        elif i == 212: #胜负（`win/lose`）
                            LoLGame_stat_data[key].append(win_dict[stats["win"]])
                        else:
                            LoLGame_stat_data[key].append(stats[key])
                    elif i <= 214: #时间轴相关键（Timeline-related keys）
                        LoLGame_stat_data[key].append(lanes[game["participants"][0]["timeline"][key]] if i == 213 else roles[game["participants"][0]["timeline"][key]])
                    else: #对局信息转换键（Keys transformed according to game information）
                        subkey = key.split("_")[0]
                        if key.endswith("_percent"): #团队占比键（Team percentage keys）
                            if i == 273: #参团率（`KP_percent`）
                                self_stat = stats["kills"] + stats["assists"]
                                total_stat = sum(map(lambda x: x["stats"]["kills"], team_participants))
                            elif i == 274: #补刀数占比（`CS_percent`）
                                self_stat = stats["totalMinionsKilled"] + stats["neutralMinionsKilled"]
                                total_stat = sum(map(lambda x: x["stats"]["totalMinionsKilled"] + x["stats"]["neutralMinionsKilled"], team_participants))
                            else:
                                self_stat = stats[subkey]
                                total_stat = sum(map(lambda x: x["stats"][subkey], team_participants))
                            value = 0 if total_stat == 0 else self_stat / total_stat
                            LoLGame_stat_data[key].append(value)
                        else: #位次键（Order keys）
                            if i == 334: #战损比位次（`KDA_order`）
                                self_stat = (stats["kills"] + stats["assists"]) / max(1, stats["deaths"])
                                stat_list = sorted(map(lambda x: (x["stats"]["kills"] + x["stats"]["assists"]) / max(1, x["stats"]["deaths"]), team_participants), reverse = True)
                            elif i == 335: #参团率位次（`KP_order`）
                                self_stat = stats["kills"] + stats["assists"]
                                stat_list = sorted(map(lambda x: x["stats"]["kills"] + x["stats"]["assists"], team_participants), reverse = True)
                            elif i == 336: #补刀数位次（`CS_order`）
                                self_stat = stats["totalMinionsKilled"] + stats["neutralMinionsKilled"]
                                stat_list = sorted(map(lambda x: x["stats"]["totalMinionsKilled"] + x["stats"]["neutralMinionsKilled"], team_participants), reverse = True)
                            elif i == 337: #伤害转化率位次（`D/G_order`）
                                self_stat = 0 if stats["goldEarned"] == 0 else stats["totalDamageDealtToChampions"] / stats["goldEarned"]
                                stat_list = sorted(map(lambda x: 0 if x["stats"]["goldEarned"] == 0 else x["stats"]["totalDamageDealtToChampions"] / x["stats"]["goldEarned"], team_participants), reverse = True)
                            elif i == 338: #金币利用率位次（`GUE_order`）
                                self_stat = 0 if stats["goldEarned"] == 0 else stats["goldSpent"] / stats["goldEarned"]
                                stat_list = sorted(map(lambda x: 0 if x["stats"]["goldEarned"] == 0 else x["stats"]["goldSpent"] / x["stats"]["goldEarned"], team_participants), reverse = True)
                            else:
                                self_stat = stats[subkey]
                                stat_list = sorted(map(lambda x: x["stats"][subkey], team_participants), reverse = True)
                            LoLGame_stat_data[key].append(0 if len(set(stat_list)) == 1 else stat_list.index(self_stat) + 1) #当所有人的数据一样时，则不用比较位次（When some stat of every player is the same, there's no need to compare it）
                if print_detail:
                    print("对局加载进度（Match loading process）：%d/%d\t对局序号（matchID）： %d" %(gameIndex + 1, len(LoLHistory["games"]["games"]), game["gameId"]))
            #因为在主函数中指定了要生成的列，所以这里没有给出键排序列表（Because the columns to print are specified in the main function, the key order list isn't given here）
            LoLGame_stat_df = pandas.DataFrame(data = LoLGame_stat_data)
            for column in LoLGame_stat_df:
                if LoLGame_stat_df[column].dtype == "bool":
                    LoLGame_stat_df[column] = LoLGame_stat_df[column].astype(str)
                    for i in range(len(LoLGame_stat_df)):
                        LoLGame_stat_df.loc[i, column] = "√" if LoLGame_stat_df[column][i] == "True" else ""
            LoLGame_stat_df = pandas.concat([pandas.DataFrame([LoLGame_stat_header])[LoLGame_stat_df.columns], LoLGame_stat_df], ignore_index = True)
            return LoLGame_stat_df
        else:
            return pandas.DataFrame(data = LoLGame_stat_header, index = [0])
    else:
        return pandas.DataFrame(data = LoLGame_stat_header, index = [0])

async def search_player_match_stats_tft(connection, puuid: str, begin: int = 0, count: int = 20) -> pandas.DataFrame:
    if verify_uuid(puuid):
        info = await get_info(connection, puuid)
        if info["info_got"]:
            #准备数据资源（Prepare data resources）
            ##云顶之弈对局记录（TFT match history）
            while True:
                TFTHistory = await (await connection.request("GET", f"/TFT-match-history/v1/products/tft/{puuid}/matches?begin={begin}&count={count}")).json()
                count = 0
                error_occurred = False
                if "errorCode" in TFTHistory:
                    if "500 Internal Server Error" in TFTHistory["message"]:
                        if not error_occurred and print_detail:
                            print("您所在大区的对局记录服务异常。尝试重新获取数据……\nThe match history service provided on your server isn't in place. Trying to recapture the history data ...")
                            error_occurred = True
                        while "errorCode" in TFTHistory and "500 Internal Server Error" in TFTHistory["message"] and count <= 3:
                            count += 1
                            if print_detail:
                                print("正在进行第%d次尝试……\nTimes trying: No. %d ..." %(count, count))
                            TFTHistory = await (await connection.request("GET", f"/TFT-match-history/v1/products/tft/{puuid}/matches?begin={begin}&count={count}")).json()
                    elif "body was empty" in TFTHistory["message"]:
                        print("这位召唤师从5月1日起就没有进行过任何云顶之弈对局。\nThis summoner hasn't played any TFT game yet since May 1st.")
                        break
                    elif "Error getting match list for summoner" in TFTHistory["message"]:
                        TFTHistory_url = "%s/lol-match-history/v1/products/tft/%s/matches?begin={begin}&count={count}" %(connection.address, puuid, begin, count)
                        print("请打开以下网址，输入如下所示的用户名和密码，打开后在命令行中按回车键继续（Please open the following website, type in the username and password accordingly and press Enter to continue）：\n网址（URL）：\t\t%s\n用户名（Username）：\triot\n密码（Password）：\t%s" %(TFTHistory_url, connection.auth_key))
                    if count > 3:
                        print("英雄联盟对局记录获取失败！请等待官方修复对局记录服务！\nTFT match history capture failure! Please wait for Tencent to fix the match history service!")
                        break
                else:
                    break
            if "errorCode" in TFTHistory:
                return pandas.DataFrame(data = TFTGame_stat_header, index = [0])
            version_re = re.compile(r"\d*\.\d*\.\d*\.\d*")
            #数据整理核心部分（Data sorting - core part）
            TFTGame_stat_data = {}
            for i in range(len(TFTGame_stat_header_keys)):
                key = TFTGame_stat_header_keys[i]
                TFTGame_stat_data[key] = []
            for gameIndex in range(len(TFTHistory["games"])):
                game = TFTHistory["games"][gameIndex]
                if game["json"]:
                    for i in range(len(TFTGame_stat_header_keys)):
                        key = TFTGame_stat_header_keys[i]
                        if i == 0: #游戏序号（`gameIndex`）
                            TFTGame_stat_data[key].append(begin + 1 + gameIndex)
                        elif i <= 10:
                            if i == 1: #对局终止情况（`endOfGameResult`）
                                if "endOfGameResult" in game["json"]:
                                    TFTGame_stat_data[key].append(endOfGameResults[game["json"][key]])
                                else:
                                    TFTGame_stat_data[key].append("")
                            elif i == 2: #对局创建时间（`gameCreation`）
                                if key in game["json"]:
                                    gameCreation = int(game["json"][key])
                                    gameCreationDate = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(gameCreation // 1000))
                                    gameCreationDate_fraction = gameCreation / 1000 - gameCreation // 1000
                                    to_append = gameCreationDate + ("{0:.3}".format(gameCreationDate_fraction))[1:5]
                                else: #14.6版本之前的云顶之弈对局信息中没有`gameCreation`这个键（The key `gameCreation` doesn't exist in information of TFT matches before Patch 14.6）
                                    to_append = ""
                                TFTGame_stat_data[key].append(to_append)
                            elif i == 3: #对局结算时间（`game_datetime`）
                                game_datetime = int(game["json"][key])
                                game_date = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(game_datetime // 1000))
                                game_date_fraction = game_datetime / 1000 - game_datetime // 1000
                                to_append = game_date + ("{0:.3}".format(game_date_fraction))[1:5]
                                TFTGame_stat_data[key].append(to_append)
                            elif i == 5: #持续时长（`game_length`）
                                TFTGame_stat_data[key].append("%d:%02d" %(int(game["json"][key]) // 60, int(game["json"][key]) % 60))
                            elif i == 6: #对局版本（`game_version`）
                                TFTGame_stat_data[key].append(version_re.search(game["json"][key]).group())
                            elif i == 8: #游戏类型（`tft_game_type`）
                                TFTGame_stat_data[key].append(gamemodes[game["json"]["queue_id"]]["description"])
                            elif i == 9: #数据版本名称（`tft_set_core_name`）
                                TFTGame_stat_data[key].append(game["json"].get(key, "")) #在云顶之弈第7赛季之前，TFTHistoryJson中无tft_set_core_name这一键（Before TFTSet7, tft_set_core_name isn't present as a key of `game["json"]`）
                            else:
                                TFTGame_stat_data[key].append(game["json"][key])
                        elif i <= 39: #对于一些容易产生争议和报错的情况，引入to_append变量以简化代码。下同（Variable `to_append` is introduced to simplify the code in case of some controversy that produces errors easily. So does the following）
                            #TFTMainPlayer = game["json"]["participants"][TFT_main_player_indices[i]]
                            for j in range(len(game["json"]["participants"])): #这里没有遵循迭代器命名原则，因为云顶之弈对局记录的赋值代码中包含了云顶之弈对局信息的赋值代码（Here the iterator naming principle isn't followed, because assignment code of TFT game information are included in those of TFT match information）
                                TFTPlayer = game["json"]["participants"][j]
                                if TFTPlayer["puuid"] == puuid:
                                    if i == 11: #玩家序号（`participantId`）
                                        TFTGame_stat_data[key].append(j + 1)
                                    elif i >= 12 and i <= 20: #强化符文相关键（Augment-related keys）
                                        if "augments" in TFTPlayer:
                                            augment_index = (i - 12) % 3
                                            subkey_index = (i - 12) // 3
                                            if augment_index < len(TFTPlayer["augments"]):
                                                TFTAugmentId = TFTPlayer["augments"][augment_index]
                                                if subkey_index == 0:
                                                    to_append = TFTAugmentId
                                                else:
                                                    to_append = TFTAugments[TFTAugmentId][key.split()[1]] if TFTAugmentId in TFTAugments else ""
                                            else:
                                                to_append = ""
                                        else:
                                            to_append = "" #云顶之弈刚出的时候，没有强化符文的概念（The concept of "augment" didn't appear at the beginning of TFT）
                                        if TFTPlayer["puuid"] == puuid:
                                            TFTGame_stat_data[key].append(to_append)
                                    elif i >= 21 and i <= 27: #小小英雄相关键（Companion-related keys）
                                        TFTCompanionId = TFTPlayer["companion"]["content_ID"]
                                        if i <= 24:
                                            to_append = TFTPlayer["companion"][key.split()[-1]]
                                        elif TFTCompanionId in TFTCompanions:
                                            to_append = TFTCompanions[TFTCompanionId][key.split()[-1]] if i <= 26 else rarities[TFTCompanions[TFTCompanionId][key.split()[-1]]]
                                        else:
                                            to_append = TFTCompanionId if i == 25 else ""
                                        if TFTPlayer["puuid"] == puuid:
                                            TFTGame_stat_data[key].append(to_append)
                                    elif i == 29: #存活回合（`last_round`）
                                        lastRound = TFTPlayer[key]
                                        if lastRound <= 3:
                                            bigRound = 1
                                            smallRound = lastRound
                                        else:
                                            bigRound = (lastRound + 3) // 7 + 1
                                            smallRound = (lastRound + 3) % 7 + 1
                                        to_append = "%d-%d" %(bigRound, smallRound)
                                        if TFTPlayer["puuid"] == puuid:
                                            TFTGame_stat_data[key].append(to_append)
                                    elif i == 34: #存活时长（`time_eliminated`）
                                        to_append = "%d:%02d" %(int(TFTPlayer[key]) // 60, int(TFTPlayer[key]) % 60)
                                        if TFTPlayer["puuid"] == puuid:
                                            TFTGame_stat_data[key].append(to_append)
                                    elif i >= 36 and i <= 39: #召唤师身份相关键（Summoner information-related keys）
                                        # if TFTPlayer["puuid"] in {"", bot_puuid}: #在云顶之弈（新手教程）中，无法通过电脑玩家的玩家通用唯一识别码（00000000-0000-0000-0000-000000000000）来查询其召唤师名称和序号（Summoner names and IDs of bot players in TFT (Tutorial) can't be searched for according to their puuid: 00000000-0000-0000-0000-000000000000）
                                        #     to_append = ""
                                        # else:
                                        #     TFTPlayer_info_recapture = 0
                                        #     TFTPlayer_info = await get_info(connection, TFTPlayer["puuid"])
                                        #     while TFTPlayer_info["network_error"] and TFTPlayer_info_recapture < 3:
                                        #         TFTPlayer_info_recapture += 1
                                        #         print("第%d/%d场对局（对局序号：%d）玩家信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of Player (puuid: %s) in Match %d / %d (matchID: %d) capture failed! Recapturing this player's information ... Times tried: %d." %(gameIndex + 1, len(TFTHistory["games"]), game["json"]["game_id"], TFTPlayer["puuid"], TFTPlayer_info_recapture, TFTPlayer["puuid"], gameIndex + 1, len(TFTHistory["games"]), game["json"]["game_id"], TFTPlayer_info_recapture))
                                        #         TFTPlayer_info = await get_info(connection, TFTPlayer["puuid"])
                                        #     if TFTPlayer_info["network_error"]:
                                        #         to_append = ""
                                        #     else:
                                        #         TFTPlayer_info_body = TFTPlayer_info["body"]
                                        #         to_append = TFTPlayer_info_body["summonerId"] if i == 36 else TFTPlayer_info_body["displayName"] if i == 37 else TFTPlayer_info_body["gameName"] if i == 38 else TFTPlayer_info_body["tagLine"]
                                        to_append = "" #这里不是为了查询其它人的战绩，因此不需要查询召唤师身份（The program is aimed at getting one user's game stats, so it's not necessary look up the other summoners' information）
                                        if TFTPlayer["puuid"] == puuid:
                                            TFTGame_stat_data[key].append(to_append)
                                    else:
                                        to_append = TFTPlayer[key]
                                        if TFTPlayer["puuid"] == puuid:
                                            TFTGame_stat_data[key].append(to_append)
                                    break
                        elif i >= 40 and i <= 130: #云顶之弈羁绊相关键（TFT trait-related keys）
                            #TFTMainPlayer_Traits = game["json"]["participants"][TFT_main_player_indices[i]]["traits"]
                            trait_index = (i - 40) // 7
                            subkey_index = (i - 40) % 7
                            for j in range(len(game["json"]["participants"])):
                                TFTPlayer = game["json"]["participants"][j]
                                TFTPlayer_Traits = TFTPlayer["traits"]
                                # if not TFTPlayer["puuid"] in {"", bot_puuid}:
                                #     TFTPlayer_info_recapture = 0
                                #     TFTPlayer_info = await get_info(connection, TFTPlayer["puuid"]) #这里的玩家信息仅用于模板羁绊的提示（The summoner information here is only used for the prompt of TemplateTrait）
                                #     while TFTPlayer_info["network_error"] and TFTPlayer_info_recapture < 3:
                                #         TFTPlayer_info_recapture += 1
                                #         print("第%d/%d场对局（对局序号：%d）玩家信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of Player (puuid: %s) in Match %d / %d (matchID: %d) capture failed! Recapturing this player's information ... Times tried: %d." %(i + 1, len(TFTHistory), game["json"]["game_id"], TFTPlayer["puuid"], TFTPlayer_info_recapture, TFTPlayer["puuid"], i + 1, len(TFTHistory), game["json"]["game_id"], TFTPlayer_info_recapture))
                                #         TFTPlayer_info = await get_info(connection, TFTPlayer["puuid"])
                                #     TFTPlayer_info_body = TFTPlayer_info["body"]
                                if trait_index < len(TFTPlayer_Traits): #在这个小于的问题上纠结了很久[敲打]——下标是从0开始的。假设API上记录了n个羁绊，那么当程序正在获取第n个羁绊时，就会引起下标越界的问题。所以这里不能使用小于等于号（I stuck at this less than sign for too long xD - note that the index begins from 0. Suppose there're totally n traits recorded in LCU API. Then, when the program is trying to capture the n-th trait, it'll throw an IndexError. That's why the "less than or equal to" sign can't be used here）
                                    TFTTrait_iter = TFTPlayer_Traits[trait_index]
                                    TFTTraitId = TFTTrait_iter["name"]
                                    if TFTTraitId == "TemplateTrait": #CommunityDragon数据库中没有收录模板羁绊的数据（Data about TemplateTrait aren't archived in CommunityDragon database）
                                        if subkey_index == 4 and not TFTPlayer["puuid"] in {"", bot_puuid}: #在艾欧尼亚的对局序号为4959597974的对局中，存在一个模板羁绊，没有tier_total这个键（There exists a TemplateTrait without the key `tier_total` in an Ionia match with matchID 4959597974）
                                            to_append = ""
                                            # print("警告：对局%d中玩家%s（玩家通用唯一识别码：%s）的第%d个羁绊是模板羁绊！\nWarning: Trait No. %d of the player %s (puuid: %s) in the match %d is TemplateTrait." %(game["json"]["game_id"], get_info_name(TFTPlayer_info_body), TFTPlayer["puuid"], trait_index + 1, trait_index + 1, get_info_name(TFTPlayer_info_body), TFTPlayer["puuid"], game["json"]["game_id"]))
                                        else:
                                            to_append = "" if subkey_index == 5 or subkey_index == 6 else TFTTrait_iter[key.split()[-1]]
                                    else:
                                        if subkey_index <= 4:
                                            to_append = traitStyles[TFTTrait_iter[key.split()[1]]] if subkey_index == 2 else TFTTrait_iter[key.split()[1]]
                                        elif TFTTraitId in TFTTraits:
                                            to_append = TFTTraits[TFTTraitId][key.split()[1]]
                                        else:
                                            to_append = ""
                                else:
                                    to_append = ""
                                if TFTPlayer["puuid"] == puuid:
                                    TFTGame_stat_data[key].append(to_append)
                        else:
                            #TFTMainPlayer_Units = game["json"]["participants"][TFT_main_player_indices[i]]["units"]
                            for j in range(len(game["json"]["participants"])):
                                TFTPlayer_Units = game["json"]["participants"][j]["units"]
                                if i <= 185: #云顶之弈英雄相关键（TFT champion-related keys）
                                    unit_index = (i - 131) // 5
                                    subkey_index = (i - 131) % 5
                                    if unit_index < len(TFTPlayer_Units):
                                        TFTChampion_iter = TFTPlayer_Units[unit_index]
                                        TFTChampionId = TFTChampion_iter["character_id"]
                                        if subkey_index >= 3:
                                            #character_id_lower = TFTPlayer_Units[unit_index]["character_id"].lower()
                                            #TFTChampion_keys_lower = list(map(lambda x: x.lower(), list(TFTChampions.keys())))
                                            if TFTChampionId in TFTChampions:
                                                to_append = TFTChampions[TFTChampionId][key.split()[-1]]
                                            elif TFTChampionId.lower() in map(lambda x: x.lower(), TFTChampions.keys()): #在获取艾欧尼亚对局序号为8390690410的英雄信息时，由于雷克塞的英雄序号大小写的原因，会引发键异常（KeyError is caused due to the case of "RekSai" string when the program is getting data from an Ionia match with matchID 8390690410）
                                                TFTChampion_index = list(map(lambda x: x.lower(), TFTChampions.keys())).index(TFTChampionId.lower())
                                                to_append = list(TFTChampions.values())[TFTChampion_index][key.split()[-1]]
                                            else:
                                                to_append = ""
                                        else:
                                            to_append = TFTPlayer_Units[unit_index][key.split()[-1]]
                                    else:
                                        to_append = ""
                                    if game["json"]["participants"][j]["puuid"] == puuid:
                                        TFTGame_stat_data[key].append(to_append)
                                else:
                                    unit_index = (i - 186) // 9
                                    item_index = (i - 186) // 3 % 3
                                    subkey_index = (i - 186) % 3
                                    if unit_index < len(TFTPlayer_Units): #很少有英雄单位可以有3个装备（Merely do champion units have full items）
                                        if "itemNames" in TFTPlayer_Units[unit_index] and item_index < len(TFTPlayer_Units[unit_index]["itemNames"]):
                                            TFTItemId = TFTPlayer_Units[unit_index]["itemNames"][item_index]
                                            if subkey_index == 0:
                                                to_append = TFTItemId
                                            elif TFTItemId in TFTItems:
                                                to_append = TFTItems[TFTItemId][key.split()[2]]
                                            elif TFTItemId in TFTAugments: #云顶之弈基础数据文件中存在部分云顶之弈装备数据文件中没有的装备（Some items are present in the TFT basic data file but absent from the TFT item data file）
                                                item_basic_dict = {"nameId": "apiName", "name": "name", "squareIconPath": "icon"} #云顶之弈装备数据文件和云顶之弈基础数据文件的格式不一致（The formats between TFT basic data and TFT item data are different）
                                                to_append = TFTAugments[TFTItemId][item_basic_dict[key.split()[2]]]
                                            else:
                                                to_append = ""
                                        elif "items" in TFTPlayer_Units[unit_index] and item_index < len(TFTPlayer_Units[unit_index]["items"]): #在12.4版本之前，装备是通过序号而不是接口名称在LCU API中被存储的（Before Patch 12.4, items are stored via itemIDs instead of itemNames）
                                            TFTItemId = TFTPlayer_Units[unit_index]["items"][item_index]
                                            if subkey_index == 0:
                                                to_append = TFTItemId
                                            elif TFTItemId in TFTItems:
                                                to_append = TFTItems[TFTItemId][key.split()[2]]
                                            elif TFTItemId in TFTAugments:
                                                item_basic_dict = {"nameId": "apiName", "name": "name", "squareIconPath": "icon"}
                                                to_append = TFTAugments[TFTItemId][item_basic_dict[key.split()[2]]]
                                            else:
                                                to_append = ""
                                        else:
                                            to_append = ""
                                    else:
                                        to_append = ""
                                    if game["json"]["participants"][j]["puuid"] == puuid:
                                        TFTGame_stat_data[key].append(to_append)
                    if print_detail:
                        print("对局加载进度（Match loading process）：%d/%d\t对局序号（matchID）： %d" %(gameIndex + 1, len(TFTHistory["games"]), game["json"]["game_id"]))
                else:
                    if print_detail:
                        print("对局加载进度（Match loading process）：%d/%d (Exceptional match neglected)" %(gameIndex + 1, len(TFTHistory["games"])))
            #因为在主函数中指定了要生成的列，所以这里没有给出键排序列表（Because the columns to print are specified in the main function, the key order list isn't given here）
            TFTGame_stat_df = pandas.DataFrame(data = TFTGame_stat_data)
            TFTGame_stat_df = pandas.concat([pandas.DataFrame([TFTGame_stat_header])[TFTGame_stat_df.columns], TFTGame_stat_df], ignore_index = True)
            return TFTGame_stat_df
        else:
            return pandas.DataFrame(data = TFTGame_stat_header, index = [0])
    else:
        return pandas.DataFrame(data = TFTGame_stat_header, index = [0])

async def Clarke_revival(connection):
    platform_config = await (await connection.request("GET", "/lol-platform-config/v1/namespaces")).json()
    platformId = platform_config["LoginDataPacket"]["platformId"]
    current_info = await (await connection.request("GET", "/lol-summoner/v1/current-summoner")).json()
    gameflow_phase = await (await connection.request("GET", "/lol-gameflow/v1/gameflow-phase")).json()
    queues_initial = await (await connection.request("GET", "/lol-game-queues/v1/queues")).json()
    queues_initial = sorted(queues_initial, key = lambda x: x["id"]) #将队列按照队列序号正序排列（Sort the queues in the ascending order of queueIds）
    queues = {queue["id"]: queue for queue in queues_initial}
    #定义获取到的玩家信息列表。元素是召唤师信息字典（Define a list of obtained summoner information）
    fetched_players = []
    #定义标记是否队友信息的字典。键是玩家通用唯一识别码，值是逻辑值（Define a dictionary to mark whether a summoner is an ally. Keys are puuids, and values are boolean values）
    ally_bool_dict = {}
    #总结性工作表呈现所有玩家的信息（Each summary sheets display all players' stats）
    LoLPlayer_stat_summary_dfs = []
    TFTPlayer_stat_summary_dfs = []
    #具体数值工作表呈现每个玩家的信息。键是召唤师名，值是数据框（Each detailed sheets display each player's stats. Keys are summonerNames, and values are dataframes）
    LoLPlayer_stat_details_dfs = {}
    TFTPlayer_stat_details_dfs = {}
    if gameflow_phase in {"ChampSelect", "InProgress", "Reconnect"}:
        #首先获取会话中的玩家信息，包括是否队友（First, get information of the players in the session, including whether they're allies or not）
        if gameflow_phase == "ChampSelect":
            champ_select_session = await (await connection.request("GET", "/lol-champ-select/v1/session")).json()
            if "errorCode" in champ_select_session:
                if champ_select_session["message"] == "No active delegate": #在没有英雄选择阶段的游戏模式中，有时gameflow_phase的结果是“ChampSelect”，但是实际上没有可用的英雄选择会话（In game modes without champ select stage, sometimes `gameflow_phase` is "ChampSelect", but there's actually no available champ select session）
                    print("英雄选择会话已过期。\nChamp select session has expired.")
                return 1
            gameflow_session = await (await connection.request("GET", "/lol-gameflow/v1/session")).json()
            if gameflow_session["gameData"]["queue"]["name"] == "":
                gamemode = gameflow_session["map"]["gameModeName"] + "(%d)" %(gameflow_session["gameData"]["queue"]["id"])
            else:
                gamemode = gameflow_session["gameData"]["queue"]["name"]
            excel_name = "Player Stats in Match %s-%s (%s).xlsx" %(platformId, champ_select_session["gameId"], gamemode)
            isSpectating = champ_select_session["isSpectating"] #从英雄选择会话信息中可以直接得到用户是否在观战（From the champ select session, it can be inferred directly whether the user is spectating）
            myTeam_puuids = list(map(lambda x: x["puuid"], champ_select_session["myTeam"]))
            for player in champ_select_session["myTeam"] + champ_select_session["theirTeam"]:
                if not player["puuid"] in {current_info["puuid"], "", bot_puuid} and (player["nameVisibilityType"] == "VISIBLE" or player["nameVisibilityType"] == ""):
                    player_info_recapture = 0
                    player_info = await get_info(connection, player["puuid"])
                    while player_info["network_error"] and player_info_recapture < 3:
                        player_info_recapture += 1
                        if print_detail:
                            print("队友信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of an player (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(player["puuid"], player_info_recapture, player["puuid"], player_info_recapture))
                        player_info = await get_info(connection, player["puuid"])
                    if player_info["network_error"] >= 3:
                        if print_detail:
                            print("队友信息（玩家通用唯一识别码：%s）获取失败！将忽略该名队友。\nInformation of an player (puuid: %s) capture failed! The program will ignore this player.")
                        continue
                    else:
                        player_info_body = player_info["body"]
                        fetched_players.append(player_info_body)
                        ally_bool_dict[player["puuid"]] = player["puuid"] in myTeam_puuids #尽管这里将位于“myTeam”中的玩家标记为队友，实际输出提示信息时，首先还是会判断是否观战（Although here all members in "myTeam" are marked allies, when the program prints prompts, it still judges whether the user is spectating）
            teamOneOnly = gameflow_session["gameData"]["queue"]["mapId"] in {"22", "30"} #在斗魂竞技场和云顶之弈中，所有玩家都归入“teamOne”中。这样就无法判断其是否是队友（In an Arena or TFT match, all players are in "teamOne", so the program can't tell whether a player is an ally）
        else:
            gameflow_session = await (await connection.request("GET", "/lol-gameflow/v1/session")).json()
            gameData = gameflow_session["gameData"]
            gamemode = gameflow_session["map"]["gameModeName"] + "(%d)" %(gameData["queue"]["id"]) if gameData["queue"]["name"] == "" else gameData["queue"]["name"]
            excel_name = "Player Stats in Match %s-%s (%s).xlsx" %(platformId, gameData["gameId"], gamemode)
            isSpectating = False #设置观战逻辑变量，确定游戏会话是不是观战的（This boolean variable is declared to tell whether the game session is spectating）
            teamOne = []
            for player in gameData["teamOne"]: #这里通过循环而不是用map函数快速获取玩家通用唯一识别码列表，是因为电脑玩家没有玩家通用唯一识别码（Here the puuid list is obtained by a loop instead of `map` function, because bot players don't have puuids）
                if "puuid" in player:
                    teamOne.append(player)
            teamTwo = []
            for player in gameData["teamTwo"]:
                if "puuid" in player:
                    teamTwo.append(player)
            if current_info["puuid"] in list(map(lambda x: x["puuid"], teamOne)): #API记录游戏中的玩家时，只会区分红蓝方，不会区分敌我。所以这里需要先判断那个阵营是我方（Players recorded in API only differentiate by blue or red team, instead of my or enemy team. So judging the own team or the enemy team is the first thing to do）
                myTeam = teamOne
                theirTeam = teamTwo
            elif current_info["puuid"] in list(map(lambda x: x["puuid"], teamTwo)):
                myTeam = teamTwo
                theirTeam = teamOne
            else:
                myTeam = teamOne + teamTwo
                theirTeam = []
                isSpectating = True
            myTeam_puuids = list(map(lambda x: x["puuid"], myTeam))
            for player in myTeam + theirTeam: #注意到这里并没有排除掉用户自己。这是为了校验程序的正确性，将自己作为空白对照（Note that here the program doesn't exclude the user itself. This is designed to verify the correctness of program execution. The user itself acts as control）
                player_info_recapture = 0
                player_info = await get_info(connection, player["puuid"])
                while player_info["network_error"] and player_info_recapture < 3:
                    player_info_recapture += 1
                    if print_detail:
                        print("玩家信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of an player (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(player["puuid"], player_info_recapture, player["puuid"], player_info_recapture))
                    player_info = await get_info(connection, player["puuid"])
                if player_info["network_error"]:
                    if print_detail:
                        print("玩家信息（玩家通用唯一识别码：%s）获取失败！将忽略该名玩家。\nInformation of an player (puuid: %s) capture failed! The program will ignore this player.")
                    continue
                else:
                    player_info_body = player_info["body"]
                    fetched_players.append(player_info_body)
                    ally_bool_dict[player["puuid"]] = player["puuid"] in myTeam_puuids
            teamOneOnly = gameData["queue"]["mapId"] in {"22", "30"} #玩家在API上的阵营划分随对局模式而不同。云顶之弈和斗魂竞技场虽然有多个阵营，但是都是记录在gameData["teamOne"]中，这需要和其它模式区分开来。该条件语句与“if gameData["queue"]["gameMode"] == "TFT" or gameData["queue"]["gameMode"] == "CHERRY"”等价，但是因为召唤师峡谷还能分成CLASSIC、URF等模式，所以这里直接用地图序号作为判断依据（The team where a player belongs varies by the game mode. Although there're actually more than 2 teams in TFT and Arena, all players are recorded in `gameData["teamOne"]`, which needs ditinguishing from other game modes. This conditional statement is equivalent to `if gameData["queue"]["gameMode"] == "TFT" or gameData["queue"]["gameMode"] == "CHERRY"`, but since there're multiple modes based on one map, like CLASSIC and URF based on Summoner's Rift, the mapId is thus taken as the judgment criterium）
    else:
        print("您目前不在英雄选择阶段或者游戏内。是否想要查看过去的英雄联盟比赛记录中某场对局中的玩家战绩？（输入任意键以查看，否则不查看。）\nYou're currently not during a champ select stage or a game. Do you want to check plater stats in a past LoL match? (Submit any non-empty string to check, or null to refuse checking.)")
        pastCheck = bool(input())
        if pastCheck:
            print('请输入对局序号。输入“0”以返回上一层。\nPlease enter the gameId. Submit "0" to return to the last step.')
            while True:
                gameId = input()
                if gameId == "0":
                    pastCheck = False
                    break
                else:
                    try:
                        gameId = int(gameId)
                    except ValueError:
                        print("请输入整数类型的对局序号！\nPlease enter the gameId of integer type!")
                    else:
                        if gameId > 0:
                            LoLGame_info = await (await connection.request("GET", f"/lol-match-history/v1/games/{gameId}")).json()
                            if "errorCode" in LoLGame_info:
                                if LoLGame_info["httpStatus"] == 404:
                                    print(f"未找到序号为{gameId}的回放文件！请重新输入。\nMatch file with matchID {gameId} not found! Please try again.")
                                else:
                                    print(LoLGame_info)
                                    print("请求失败！请切换一个对局序号或稍后重试。\nRequest failed! Please change a gameId or try it again later.")
                            else:
                                break
                        else:
                            print("请输入一个正整数！\nPlease enter a positive integer.")
        if pastCheck:
            gamemode = queues[LoLGame_info["queueId"]]["name"] if LoLGame_info["queueId"] in queues else LoLGame_info["gameMode"] + "(%d)" %(LoLGame_info["queueId"])
            excel_name = "Player Stats in Match %s-%s (%s).xlsx" %(platformId, gameId, gamemode)
            isSpectating = False #严格地讲并不算观战，但其作用与其它地方该变量的作用相同（Seriously, this situation can't be regarded as spectating. However, it plays the same role as in other places）
            teamOne = []
            teamTwo = []
            players = {participant["participantId"]: participant["player"] for participant in LoLGame_info["participantIdentities"]}
            for player in LoLGame_info["participants"]:
                if player["teamId"] == 100:
                    teamOne.append(players[player["participantId"]])
                elif player["teamId"] == 200:
                    teamTwo.append(players[player["participantId"]])
            if current_info["puuid"] in list(map(lambda x: x["puuid"], teamOne)):
                myTeam = teamOne
                theirTeam = teamTwo
            elif current_info["puuid"] in list(map(lambda x: x["puuid"], teamTwo)):
                myTeam = teamTwo
                theirTeam = teamOne
            else:
                myTeam = teamOne + teamTwo
                theirTeam = []
                isSpectating = True
            myTeam_puuids = list(map(lambda x: x["puuid"], myTeam))
            for player in myTeam + theirTeam:
                if not player["puuid"] in {"", bot_puuid}:
                    player["displayName"] = player["summonerName"] #对局信息中的显示名是“summonerName”，而不是“displayName”，为了避免后面`get_info_name`函数提示格式错误，所以临时添加了“displayName”键（The displayName key is names as "summonerName" instead of "displayName" in the match information, so to avoid the format error from `get_info_name` function, "displayName" key is temporarily added）
                    fetched_players.append(player)
                    ally_bool_dict[player["puuid"]] = player["puuid"] in myTeam_puuids
            teamOneOnly = LoLGame_info["mapId"] in {"22", "30"}
    if len(fetched_players) == 0:
        if gameflow_phase in {"ChampSelect", "InProgress", "Reconnect"}:
            print("未在当前对局中发现其它人类玩家。\nThere's not any other human player in this game.")
        else:
            print("请确保您正处于英雄选择阶段或者游戏内。\nPlease confirm you're during a champ select stage or a game.")
    else:
        social_leaderboard_df = await sort_social_leaderboard(connection, list(map(lambda x: x["puuid"], fetched_players)))
        social_leaderboard_statistics_output_order = [11, 1, 2, 3, 0, 16, 4, 17, 15, 7, 5, 9, 10, 8, 19]
        social_leaderboard_fields = list(map(lambda x: social_leaderboard_header_keys[x], social_leaderboard_statistics_output_order))
        social_leaderboard_df_export = social_leaderboard_df.loc[:, social_leaderboard_fields]
        print("请输入您想要查询的对局数量，默认为最近20场。最大200场。\nPlease enter the number of matches you want to search, 20 by default and 200 at maximum.")
        while True:
            count = input()
            if count == "":
                count = 20
                break
            else:
                try:
                    count = int(count)
                except ValueError:
                    print("请输入一个整数！\nPlease enter an integer!")
                else:
                    if count > 0:
                        break
                    else:
                        print("请输入一个正整数！\nPlease enter a positive integer!")
        queues_df = await get_queue_data(connection)
        queues_df_fields_to_print = ["id", "name", "mapId", "category", "pickMode"]
        queues_df_indices_to_select = [0] + list(queues_df[(queues_df["queueAvailability"] == "√") | (queues_df["isVisible"] == "√")].index)
        print("是否需要对游戏队列取子集？（输入任意键以开始打草稿，否则直接开始输入队列序号。）\nDo you want to get a subset of the current game mode data? (Submit any non-empty string to make a draft, or null to input the queue ids directly.)")
        draft = bool(input())
        if draft:
            print("请选择草稿选项：\nPlease select a draft option:\n0\t退出草稿（Quit drafting）\n1\t迭代式取子集（Take subsets iteratively）")
            while True:
                draft_option = input()
                if draft_option == "":
                    continue
                elif draft_option[0] == "0":
                    break
                elif draft_option[0] == "1":
                    scope = {"format_df": format_df, "df": queues_df.copy(deep = True), "queues": queues_initial, "fields": queues_df_fields_to_print}
                    print('示例（Examples）：\nprint(dir())\nprint(format_df(df[(df["type"] == "URF") & (df["assetMutator"] == "PICKURF")].loc[1:, fields])[0])\n输入“-1”以退出取子集。\nSubmit "-1" to quit taking subsets.')
                    subscope(scope)
                else:
                    print("您的输入有误！请重新输入。\nERROR input! Please try again.")
                print("请选择草稿选项：\nPlease select a draft option:\n0\t退出草稿（Quit drafting）\n1\t迭代式取子集（Take subsets iteratively）")
        print("请输入队列序号以选择游戏模式。直接按回车键以显示所有对局。\nPlease select a game mode by entering queue ids. Press Enter directly to display all matches.")
        print(format_df(queues_df.loc[queues_df_indices_to_select, queues_df_fields_to_print])[0])
        print('''变量提示（Variable hint）：\nqueues_df = await get_queue_data(connection)\n示例（Examples）：\n\nall\n420\n[420, 440]\n#筛选所有英雄联盟队列（Select all LoL queues）\nlist(queues_df[queues_df["mapId"] != 22].loc[1:, "id"])\n#筛选所有云顶之弈队列（Select all TFT queues）\nlist(queues_df[queues_df["mapId"] == 22].loc[:, "id"])\n#筛选召唤师峡谷的排位队列（Select ranked queues in Summoner's Rift）\nlist(queues_df[(queues_df["type"].str.startswith("RANKED")) & (queues_df["mapId"] == 11)].loc[:, "id"])\n输入以“#”开头的序号以使用本程序提供的快捷选项：\nEnter a number starting with "#" to choose a quick option (if any one fits your demand):''')
        for option in queueId_options:
            print("%s\t%s\n\t%s" %(option, queueId_options[option]["description"], queueId_options[option]["expression"]))
        queue_select = True #决定是否在后续输出时对对局的游戏模式进行筛选（Decides whether to select the game modes among the matches subsequently）
        while True:
            queueId_str = input()
            if queueId_str == "" or queueId_str == "all":
                queue_select = False
                break
            elif queueId_str == "0":
                return 1
            elif queueId_str in queueId_options:
                queueIds = eval(queueId_options[queueId_str]["expression"])
                break
            else:
                try:
                    queueIds = eval(queueId_str)
                except:
                    print("您的输入有误！请重新输入。\nERROR input! Please try again.")
                else:
                    if isinstance(queueIds, int) and queueIds in list(queues_df.loc[:, "id"]):
                        queueIds = [queueIds]
                        break
                    elif isinstance(queueIds, list) and all(map(lambda x: isinstance(x, int) and (x in {-1, 0} or x in list(queues_df.loc[1:, "id"])), queueIds)): #自定义对局的对局序号是0（QueueId of a custom game is 0）
                        break
                    else:
                        print("您的输入有误！请重新输入。\nERROR input! Please try again.")
        search_LoL = search_TFT = False #通过用户选择的队列序号来决定查看英雄联盟还是云顶之弈的数据（Decide to check LoL or TFT game stats through the queueIds selected by the user）
        if queue_select:
            for queueId in queueIds:
                if search_LoL and search_TFT: #当从用户输入的队列序号中同时检测到了英雄联盟和云顶之弈队列，就不用再检测了（When the program detects both LoL and TFT game queues, there's no need to detect further）
                    break
                if queueId in queues:
                    if queues[queueId]["gameMode"] == "TFT":
                        search_TFT = True
                    else:
                        search_LoL = True
                elif queueId in {-1, 0}:
                    search_LoL = True
        else:
            search_LoL = search_TFT = True
        LoLGame_stat_fields_summary = ["gameModeName", "champion_name", "K/D/A", "CS", "gameDuration", "win/lose", "KP_percent", "goldEarned", "KDA", "totalDamageDealtToChampions", "totalDamageTaken", "totalHeal", "KP_order", "goldEarned_order", "KDA_order", "totalDamageDealtToChampions_order", "totalDamageTaken_order", "totalHeal_order"]
        LoLPlayer_stat_summary_dfs.append(pandas.concat([pandas.DataFrame(data = {"summonerName": "召唤师名"}, index = [0]), pandas.DataFrame(data = LoLGame_stat_header, index = [0]).loc[:, LoLGame_stat_fields_summary]], axis = 1))
        LoLGame_stat_statistics_output_order = [0, 5, 3, 4, 10, 6, 11, 9, 8, 27, 30, 125, 31, 32, 34, 35, 40, 213, 214, 151, 152, 153, 154, 155, 156, 157, 187, 199, 188, 200, 189, 201, 190, 202, 191, 203, 192, 204, 67, 45, 38, 205, 206, 207, 210, 211, 41, 136, 137, 69, 66, 70, 49, 48, 53, 52, 51, 50, 46, 140, 126, 79, 145, 130, 138, 132, 107, 73, 142, 131, 106, 72, 141, 68, 43, 42, 134, 139, 133, 108, 74, 143, 44, 146, 149, 148, 127, 147, 56, 208, 57, 209, 135, 75, 77, 76, 144, 58, 71, 183, 185, 171, 165, 172, 166, 173, 167, 174, 168, 175, 169, 176, 170, 39, 47, 129, 54, 55, 212, 128, 226, 220, 215, 273, 216, 260, 228, 225, 229, 221, 263, 252, 238, 268, 254, 261, 256, 240, 232, 265, 255, 239, 231, 264, 227, 218, 217, 258, 262, 257, 241, 233, 266, 219, 269, 272, 271, 253, 270, 222, 223, 259, 234, 236, 235, 274, 267, 224, 230, 276, 287, 281, 275, 334, 335, 337, 277, 321, 289, 286, 290, 282, 324, 313, 299, 329, 315, 322, 317, 301, 293, 326, 316, 300, 292, 325, 288, 279, 278, 319, 323, 318, 302, 294, 327, 280, 330, 333, 332, 314, 331, 283, 284, 338, 320, 295, 296, 297, 336, 328, 285, 291]
        LoLGame_stat_fields_details = list(map(lambda x: LoLGame_stat_header_keys[x], LoLGame_stat_statistics_output_order))
        #LoLGame_stat_fields_details_to_print = ["gameIndex", "gameCreationDate", "gameModeName", "champion_name", "champion_alias", "K/D/A", "KDA", "win/lose"]
        TFTGame_stat_fields_summary = ["tft_game_type", "last_round", "total_damage_to_players", "players_eliminated", "placement"]
        TFTPlayer_stat_summary_dfs.append(pandas.concat([pandas.DataFrame(data = {"summonerName": "召唤师名"}, index = [0]), pandas.DataFrame(data = TFTGame_stat_header, index = [0]).loc[:, TFTGame_stat_fields_summary]], axis = 1))
        TFTGame_stat_statistics_output_order = [0, 4, 2, 3, 5, 7, 8, 6, 10, 25, 26, 27, 30, 29, 34, 28, 35, 32, 31, 15, 16, 17, 134, 132, 133, 187, 190, 193, 139, 137, 138, 196, 199, 202, 144, 142, 143, 205, 208, 211, 149, 147, 148, 214, 217, 220, 154, 152, 153, 223, 226, 229, 159, 157, 158, 232, 235, 238, 164, 162, 163, 241, 244, 247, 169, 167, 168, 250, 253, 256, 174, 172, 173, 259, 262, 265, 179, 177, 178, 268, 271, 274, 184, 182, 183, 277, 280, 283, 45, 41, 42, 43, 44, 52, 48, 49, 50, 51, 59, 55, 56, 57, 58, 66, 62, 63, 64, 65, 73, 69, 70, 71, 72, 80, 76, 77, 78, 79, 87, 83, 84, 85, 86, 94, 90, 91, 92, 93, 101, 97, 98, 99, 100, 108, 104, 105, 106, 107, 115, 111, 112, 113, 114, 122, 118, 119, 120, 121, 129, 125, 126, 127, 128]
        TFTGame_stat_fields_details = list(map(lambda x: TFTGame_stat_header_keys[x], TFTGame_stat_statistics_output_order))
        #TFTGame_stat_fields_details_to_print = ["gameIndex", "game_datetime", "tft_game_type", "companion name", "last_round", "placement"]
        max_numPlayersPerTeam_lol = 0 #仅用于后续套用条件格式时确定三色刻度的上限。在大多数情况下，这个变量的值是5（Only used to determine the maximum value of the 3-Color Scale conditional formatting. In most cases, the value of this variable is 5）
        for i in range(len(fetched_players)):
            player = fetched_players[i]
            player_summonerName = get_info_name(player)
            if search_LoL:
                print("[%d/%d]" %((search_LoL + search_TFT) * i + 1, (search_LoL + search_TFT) * len(fetched_players)), end = "")
                print(f"正在获取{player_summonerName}的最近{count}场英雄联盟对局记录……\nLoading the recent {count} LoL match(es) of {player_summonerName} ...")
                LoLGame_stat_df = await search_player_match_stats_lol(connection, player["puuid"], begIndex = 0, endIndex = count - 1)
                #print(f"{player_summonerName}的最近{count}场英雄联盟对局记录获取完成。\nThe recent {count} LoL match(es) of {player_summonerName} have/has been loaded.")
                if queue_select:
                    LoLGame_stat_indices_to_select = [0] + list(LoLGame_stat_df[LoLGame_stat_df["queueId"].isin(queueIds)].index)
                    LoLGame_stat_df = LoLGame_stat_df.loc[LoLGame_stat_indices_to_select, :]
                max_numPlayersPerTeam_lol = 5 if queue_select and len(LoLGame_stat_indices_to_select) == 1 or len(LoLGame_stat_df) == 1 else max(max_numPlayersPerTeam_lol, max(map(lambda x: 5 if x == 0 else 2 if queues[x]["gameMode"] == "CHERRY" else queues[x]["numPlayersPerTeam"], LoLGame_stat_df.loc[1:, "queueId"]))) #自定义对局的队伍规模视为5；斗魂竞技场的队伍规模虽然在API中记录为16，但这里应该考虑的是子阵营（The team size of any custom game is regarded as 5; although the team size of an Arena game is recorded as in LCU API, the subteam has more reference value）
                #print(LoLGame_stat_df.loc[:, LoLGame_stat_fields_details])
                LoLPlayer_stat_details_dfs[player_summonerName] = LoLGame_stat_df.loc[:, LoLGame_stat_fields_details]
                LoLPlayer_stat_summary_df = pandas.concat([pandas.DataFrame(data = {"summonerName": ["召唤师名"] + [player_summonerName] * (len(LoLGame_stat_df) - 1)}, index = LoLGame_stat_df.index), LoLGame_stat_df.loc[:, LoLGame_stat_fields_summary]], axis = 1)
                LoLPlayer_stat_summary_dfs.append(LoLPlayer_stat_summary_df.loc[1:, :])
            if search_TFT:
                print("[%d/%d]" %((search_LoL + search_TFT) * i + 1 + search_LoL, (search_LoL + search_TFT) * len(fetched_players)), end = "")
                print(f"正在获取{player_summonerName}的最近{count}场云顶之弈对局记录……\nLoading the recent {count} TFT match(es) of {player_summonerName} ...")
                TFTGame_stat_df = await search_player_match_stats_tft(connection, player["puuid"], begin = 0, count = count - 1)
                #print(f"{player_summonerName}的最近{count}场云顶之弈对局记录获取完成。\nThe recent {count} TFT match(es) of {player_summonerName} have/has been loaded.")
                if queue_select:
                    TFTGame_stat_indices_to_select = [0] + list(TFTGame_stat_df[TFTGame_stat_df["queue_id"].isin(queueIds)].index)
                    TFTGame_stat_df = TFTGame_stat_df.loc[TFTGame_stat_indices_to_select, :]
                TFTPlayer_stat_details_dfs[player_summonerName] = TFTGame_stat_df.loc[:, TFTGame_stat_fields_details]
                TFTPlayer_stat_summary_df = pandas.concat([pandas.DataFrame(data = {"summonerName": ["召唤师名"] + [player_summonerName] * (len(TFTGame_stat_df) - 1)}, index = TFTGame_stat_df.index), TFTGame_stat_df.loc[:, TFTGame_stat_fields_summary]], axis = 1)
                TFTPlayer_stat_summary_dfs.append(TFTPlayer_stat_summary_df.loc[1:, :])
                #print(TFTGame_stat_df.loc[:, TFTGame_stat_fields_details])
        sheet_headers = {get_info_name(player): "" if isSpectating or teamOneOnly else "Ally - " if ally_bool_dict[player["puuid"]] else "Enemy - " for player in fetched_players}
        if search_LoL:
            LoLPlayers_summary_df = pandas.concat(LoLPlayer_stat_summary_dfs, ignore_index = True)
        if search_TFT:
            TFTPlayers_summary_df = pandas.concat(TFTPlayer_stat_summary_dfs, ignore_index = True)
        #设置输出到工作簿时的格式（Set the format when the dataframe is exported to the workbook）
        if search_LoL:
            twoDigitPercentage_columns_lol_summary = ["KP_percent"] #百分比（Percentage）
            oneDigitFloat_columns_lol_summary = ["KDA"] #一位小数（One-digit float）
            colorScale_columns_lol_summary = [column for column in LoLGame_stat_fields_summary if column.endswith("_order")] #条件格式——渐变颜色（Conditional formatting - color scaling）
            dataBar_columns_lol_summary = [column for column in LoLGame_stat_fields_summary if column.endswith("_percent")] #条件格式——数据条（Conditional formatting - data bar）
            twoDigitPercentage_columns_lol_details = [column for column in LoLGame_stat_fields_details if column.endswith("_percent") or column == "GUE"] #百分比（Percentage）
            oneDigitFloat_columns_lol_details = ["KDA"] #一位小数（One-digit float）
            threeDigitFloat_columns_lol_details = ["CSPM", "D/G", "GPM"] #三位小数（Three-digit float）
            colorScale_columns_lol_details = [column for column in LoLGame_stat_fields_details if column.endswith("_order")] #条件格式——渐变颜色（Conditional formatting - color scaling）
            dataBar_columns_lol_details = [column for column in LoLGame_stat_fields_details if column.endswith("_percent")] #条件格式——数据条（Conditional formatting - data bar）
            order_colorScaleRule_lol = ColorScaleRule(start_type = "num", start_value = 1, start_color = "63BE7B", mid_type = "percentile", mid_value = 50, mid_color = "FFEB84", end_type = "num", end_value = max_numPlayersPerTeam_lol, end_color = "FF6B6B") #跳过名次为0的单元格。这里`end_value`的选取可以讨论一下，可以选取所有对局的队列信息中记录的队伍规模的最大值（Skip the order cells whose values are 0. Here the value of `end_value` is worth discussion: it may take the maximum of `numPlayersPerTeam` recorded in the queue data of the corresponding queueIds）
            percent_dataBarRule_lol = DataBarRule(start_type = "percentile", start_value = 0, end_type = "percentile", end_value = 100, color = Color("008AEF"), minLength = None, maxLength = None)
        #导出玩家战绩（Export player stats）
        print("正在保存玩家战绩……\nSaving player stats ...")
        while True:
            try:
                with pandas.ExcelWriter(path = excel_name, engine = "openpyxl") as writer: #使用openpyxl引擎套用条件格式（Use "openpyxl" engine to add conditional formats）
                    social_leaderboard_df_export.to_excel(excel_writer = writer, sheet_name = "Social Leaderboard")
                    if print_detail:
                        print("玩家排位信息已汇总。\nSocial leaderboard has been summarized.")
                    if search_LoL:
                        LoLPlayers_summary_df.to_excel(excel_writer = writer, sheet_name = "Player Summary (LoL)")
                        worksheet = writer.sheets["Player Summary (LoL)"]
                        worksheet.conditional_formatting.rules = [] #读取时清空原规则（Clear original rules when reading）
                        #套用保留两位小数的百分比格式（Two-digit percentage）
                        for column in twoDigitPercentage_columns_lol_summary:
                            col_idx = LoLPlayers_summary_df.columns.get_loc(column) + 2 #Excel中的第一列（A列）的索引是1，且又是数据框的索引列【The index of the first column (Column A) in Excel is 1, and this column is the index of column of the dataframe）
                            for row in range(3, len(LoLPlayers_summary_df) + 2):
                                worksheet.cell(row = row, column = col_idx).number_format = numbers.FORMAT_PERCENTAGE_00
                        #套用一位小数（One-digit float）
                        for column in oneDigitFloat_columns_lol_summary:
                            col_idx = LoLPlayers_summary_df.columns.get_loc(column) + 2
                            for row in range(3, len(LoLPlayers_summary_df) + 2):
                                worksheet.cell(row = row, column = col_idx).number_format = "0.0"
                        #胜负颜色（Win/Lose color）
                        col_idx = LoLPlayers_summary_df.columns.get_loc("win/lose") + 2
                        col_letter = get_column_letter(col_idx)
                        rangeStr = "%s3:%s%d" %(col_letter, col_letter, len(LoLPlayers_summary_df) + 2)
                        win_formulaRule_lol = FormulaRule(formula = ['$%s3="%s"' %(col_letter, win_dict[True])], stopIfTrue = True, fill = PatternFill(start_color = "63BE7B", end_color = "63BE7B", fill_type = "solid"))
                        lose_formulaRule_lol = FormulaRule(formula = ['$%s3="%s"' %(col_letter, win_dict[False])], stopIfTrue = True, fill = PatternFill(start_color = "FF6B6B", end_color = "FF6B6B", fill_type = "solid"))
                        worksheet.conditional_formatting.add(rangeStr, win_formulaRule_lol)
                        worksheet.conditional_formatting.add(rangeStr, lose_formulaRule_lol)
                        #百分比颜色（Percent color）
                        rangeStrs = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
                        for i in range(len(dataBar_columns_lol_summary)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
                            column = dataBar_columns_lol_summary[i]
                            if i == 0:
                                startCol_idx = endCol_idx = LoLPlayers_summary_df.columns.get_loc(column) + 2
                            else:
                                col_idx = LoLPlayers_summary_df.columns.get_loc(column) + 2
                                if col_idx == endCol_idx + 1: #如果下一个要添加条件格式的列号与上一个要添加条件格式的列号差1，那么这两列是相邻的，即连贯的（If the number of the current column to add conditional format is greater than the number of the predecessive column to add conditional format by 1, then these two columns are continuous）
                                    endCol_idx = col_idx
                                else: #如果两列不相邻，则提取得到上一个连贯的区域（If these two columns aren't continuous, then get the previous continuous area）
                                    startCol_letter = get_column_letter(startCol_idx)
                                    endCol_letter = get_column_letter(endCol_idx)
                                    rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLPlayers_summary_df) + 2)
                                    rangeStrs.append(rangeStr)
                                    startCol_idx = endCol_idx = col_idx #将区域的起始列和终止列设置为当前列（Set the starting and ending columns as the current column）
                        else: #执行完成后，把最后一个连贯区域也加上（After the for-loop finishes, add the last continuous area）
                            startCol_letter = get_column_letter(startCol_idx)
                            endCol_letter = get_column_letter(endCol_idx)
                            rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLPlayers_summary_df) + 2)
                            rangeStrs.append(rangeStr)
                        for rangeStr in rangeStrs:
                            worksheet.conditional_formatting.add(rangeStr, percent_dataBarRule_lol)
                        #位次颜色（Order color）
                        rangeStrs = []
                        rangeTuples = []
                        for i in range(len(colorScale_columns_lol_summary)):
                            column = colorScale_columns_lol_summary[i]
                            if i == 0:
                                startCol_idx = endCol_idx = LoLPlayers_summary_df.columns.get_loc(column) + 2
                            else:
                                col_idx = LoLPlayers_summary_df.columns.get_loc(column) + 2
                                if col_idx == endCol_idx + 1:
                                    endCol_idx = col_idx
                                else:
                                    startCol_letter = get_column_letter(startCol_idx)
                                    endCol_letter = get_column_letter(endCol_idx)
                                    rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLPlayers_summary_df) + 2)
                                    rangeStrs.append(rangeStr)
                                    rangeTuples.append((startCol_letter, endCol_letter))
                                    startCol_idx = endCol_idx = col_idx
                        else:
                            startCol_letter = get_column_letter(startCol_idx)
                            endCol_letter = get_column_letter(endCol_idx)
                            rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLPlayers_summary_df) + 2)
                            rangeStrs.append(rangeStr)
                            rangeTuples.append((startCol_letter, endCol_letter))
                        for i in range(len(rangeStrs)):
                            rangeStr = rangeStrs[i]
                            rangeTuple = rangeTuples[i]
                            order_noFillRule = FormulaRule(formula = ["%s3=0" %(rangeTuple[0])], stopIfTrue = True, fill = PatternFill(fill_type = None))
                            worksheet.conditional_formatting.add(rangeStr, order_noFillRule)
                            worksheet.conditional_formatting.add(rangeStr, order_colorScaleRule_lol)
                        if print_detail:
                            print("英雄联盟战绩已汇总。\nLoL game stats have been summarized.")
                    if search_TFT:
                        TFTPlayers_summary_df.to_excel(excel_writer = writer, sheet_name = "Player Summary (TFT)")
                        if print_detail:
                            print("云顶之弈战绩已汇总。\nTFT game stats have been summarized.")
                    if search_LoL:
                        for summonerName in LoLPlayer_stat_details_dfs:
                            LoLPlayer_stat_details_df = LoLPlayer_stat_details_dfs[summonerName]
                            LoLPlayer_stat_details_df.to_excel(excel_writer = writer, sheet_name = "%s%s (LoL)" %(sheet_headers[summonerName], summonerName))
                            worksheet = writer.sheets["%s%s (LoL)" %(sheet_headers[summonerName], summonerName)]
                            worksheet.conditional_formatting.rules = [] #读取时清空原规则（Clear original rules when reading）
                            #套用保留两位小数的百分比格式（Two-digit percentage）
                            for column in twoDigitPercentage_columns_lol_details:
                                col_idx = LoLPlayer_stat_details_df.columns.get_loc(column) + 2
                                for row in range(3, len(LoLPlayer_stat_details_df) + 2):
                                    worksheet.cell(row = row, column = col_idx).number_format = numbers.FORMAT_PERCENTAGE_00
                            #套用一位小数（One-digit float）
                            for column in oneDigitFloat_columns_lol_details:
                                col_idx = LoLPlayer_stat_details_df.columns.get_loc(column) + 2
                                for row in range(3, len(LoLPlayer_stat_details_df) + 2):
                                    worksheet.cell(row = row, column = col_idx).number_format = "0.0"
                            #套用三位小数（One-digit float）
                            for column in threeDigitFloat_columns_lol_details:
                                col_idx = LoLPlayer_stat_details_df.columns.get_loc(column) + 2
                                for row in range(3, len(LoLPlayer_stat_details_df) + 2):
                                    worksheet.cell(row = row, column = col_idx).number_format = "0.000"
                            #胜负颜色（Win/Lose color）
                            col_idx = LoLPlayer_stat_details_df.columns.get_loc("win/lose") + 2
                            col_letter = get_column_letter(col_idx)
                            rangeStr = "%s3:%s%d" %(col_letter, col_letter, len(LoLPlayer_stat_details_df) + 2)
                            win_formulaRule_lol = FormulaRule(formula = ['$%s3="%s"' %(col_letter, win_dict[True])], stopIfTrue = True, fill = PatternFill(start_color = "63BE7B", end_color = "63BE7B", fill_type = "solid"))
                            lose_formulaRule_lol = FormulaRule(formula = ['$%s3="%s"' %(col_letter, win_dict[False])], stopIfTrue = True, fill = PatternFill(start_color = "FF6B6B", end_color = "FF6B6B", fill_type = "solid"))
                            worksheet.conditional_formatting.add(rangeStr, win_formulaRule_lol)
                            worksheet.conditional_formatting.add(rangeStr, lose_formulaRule_lol)
                            #百分比颜色（Percent color）
                            rangeStrs = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
                            for i in range(len(dataBar_columns_lol_details)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
                                column = dataBar_columns_lol_details[i]
                                if i == 0:
                                    startCol_idx = endCol_idx = LoLPlayer_stat_details_df.columns.get_loc(column) + 2
                                else:
                                    col_idx = LoLPlayer_stat_details_df.columns.get_loc(column) + 2
                                    if col_idx == endCol_idx + 1: #如果下一个要添加条件格式的列号与上一个要添加条件格式的列号差1，那么这两列是相邻的，即连贯的（If the number of the current column to add conditional format is greater than the number of the predecessive column to add conditional format by 1, then these two columns are continuous）
                                        endCol_idx = col_idx
                                    else: #如果两列不相邻，则提取得到上一个连贯的区域（If these two columns aren't continuous, then get the previous continuous area）
                                        startCol_letter = get_column_letter(startCol_idx)
                                        endCol_letter = get_column_letter(endCol_idx)
                                        rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLPlayer_stat_details_df) + 2)
                                        rangeStrs.append(rangeStr)
                                        startCol_idx = endCol_idx = col_idx #将区域的起始列和终止列设置为当前列（Set the starting and ending columns as the current column）
                            else: #执行完成后，把最后一个连贯区域也加上（After the for-loop finishes, add the last continuous area）
                                startCol_letter = get_column_letter(startCol_idx)
                                endCol_letter = get_column_letter(endCol_idx)
                                rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLPlayer_stat_details_df) + 2)
                                rangeStrs.append(rangeStr)
                            for rangeStr in rangeStrs:
                                worksheet.conditional_formatting.add(rangeStr, percent_dataBarRule_lol)
                            #位次颜色（Order color）
                            rangeStrs = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
                            rangeTuples = []
                            for i in range(len(colorScale_columns_lol_details)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
                                column = colorScale_columns_lol_details[i]
                                if i == 0:
                                    startCol_idx = endCol_idx = LoLPlayer_stat_details_df.columns.get_loc(column) + 2
                                else:
                                    col_idx = LoLPlayer_stat_details_df.columns.get_loc(column) + 2
                                    if col_idx == endCol_idx + 1: #如果下一个要添加条件格式的列号与上一个要添加条件格式的列号差1，那么这两列是相邻的，即连贯的（If the number of the current column to add conditional format is greater than the number of the predecessive column to add conditional format by 1, then these two columns are continuous）
                                        endCol_idx = col_idx
                                    else: #如果两列不相邻，则提取得到上一个连贯的区域（If these two columns aren't continuous, then get the previous continuous area）
                                        startCol_letter = get_column_letter(startCol_idx)
                                        endCol_letter = get_column_letter(endCol_idx)
                                        rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLPlayer_stat_details_df) + 2)
                                        rangeStrs.append(rangeStr)
                                        rangeTuples.append((startCol_letter, endCol_letter))
                                        startCol_idx = endCol_idx = col_idx #将区域的起始列和终止列设置为当前列（Set the starting and ending columns as the current column）
                            else: #执行完成后，把最后一个连贯区域也加上（After the for-loop finishes, add the last continuous area）
                                startCol_letter = get_column_letter(startCol_idx)
                                endCol_letter = get_column_letter(endCol_idx)
                                rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLPlayer_stat_details_df) + 2)
                                rangeStrs.append(rangeStr)
                                rangeTuples.append((startCol_letter, endCol_letter))
                            for i in range(len(rangeStrs)):
                                rangeStr = rangeStrs[i]
                                rangeTuple = rangeTuples[i]
                                order_noFillRule = FormulaRule(formula = ["%s3=0" %(rangeTuple[0])], stopIfTrue = True, fill = PatternFill(fill_type = None))
                                worksheet.conditional_formatting.add(rangeStr, order_noFillRule)
                                worksheet.conditional_formatting.add(rangeStr, order_colorScaleRule_lol)
                            if print_detail:
                                print(f"{summonerName}的英雄联盟详细战绩已导出。\n{summonerName}'s detailed LoL game stats have been exported.")
                    if search_TFT:
                        for summonerName in TFTPlayer_stat_details_dfs:
                            TFTPlayer_stat_details_dfs[summonerName].to_excel(excel_writer = writer, sheet_name = "%s%s (TFT)" %(sheet_headers[summonerName], summonerName))
                            if print_detail:
                                print(f"{summonerName}的云顶之弈详细战绩已导出。\n{summonerName}'s detailed TFT game stats have been exported.")
                if gameflow_phase == "ChampSelect":
                    print(f'英雄选择阶段的玩家战绩已导出。请查看同目录下的“{excel_name}”。\nPlayer game stats during champ select have been exported. Please check "{excel_name}" under the same folder.')
                else:
                    print(f'本场对局的玩家战绩已导出。请查看同目录下的“{excel_name}”。\nPlayer game stats in this game have been exported. Please check "{excel_name}" under the same folder.')
            except PermissionError:
                print("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                input()
            #由于上面是覆盖写而不是追加写，而且是在脚本运行目录下生成文件，所以不会出现文件不存在的报错（Because the above writing style is overwriting instead of appending, and the generated file is under the same folder as this program, FileNotFoundError shouldn't happen）
            else:
                if args.open_after_save:
                    open_action = True
                else:
                    print("输入任意非空字符串以打开该文件，或者直接进行下一步。\nSubmit any non-empty string to open this file, or null to begin the next search.")
                    open_action = bool(input())
                if open_action:
                    #os.system(f'"{excel_name}"') #这个命令有时会引起主程序卡顿（Sometimes this command make the main program stuck）
                    excel_filePath = os.path.join(wd, excel_name)
                    process = subprocess.Popen(f'start "" "{excel_filePath}"', shell = True) #不显示新的窗口（New window won't be displayed）
                    created_processes.append(process.pid)
                break

#-----------------------------------------------------------------------------
# websocket
#-----------------------------------------------------------------------------
@connector.ready
async def connect(connection):
    await get_summoner_data(connection)
    await define_global_variables(connection)
    global print_detail
    if args.verbose:
        print_detail = True
    elif args.nonverbose:
        print_detail = False
    else:
        print("是否输出具体加载进度？（输入任意键以输出具体过程，否则输出简略过程。）\nDo you want the program to output the loading process in details? (Submit any non-empty string to output the detailed process, or null to output the simplified process.)")
        print_detail = bool(input())
    while True:
        print('按回车键以继续，或者输入任意非空字符串以退出程序。\nPress Enter to continue, or submit any non-empty string to exit the program.')
        if input() != "":
            break
        await Clarke_revival(connection)
    # process_df = check_proc_trees(created_processes)
    # if len(process_df) > 1:
    #     print("是否清理残留进程？注意，这会关闭已经打开的工作簿。（输入任意键以清理，否则不清理。）\nDo you want to clear the remaining processes? Note that the workbook process will also be killed. (Submit any non-empty string to clear, or null to deny.)\n本程序创建的进程如下：\nProcesses created by this program are listed below:")
    #     print(format_df(process_df)[0])
    #     clear_process = bool(input())
    #     if clear_process:
    #         for i in range(1, len(process_df)):
    #             pid = process_df.at[i, "pid"]
    #             pName = process_df.at[i, "name"]
    #             try:
    #                 process = psutil.Process(pid)
    #             except psutil.NoSuchProcess:
    #                 if print_detail:
    #                     print(f"未找到进程{pName}（{pid}）。将跳过该进程。\nProcess {pName} ({pid}) not found! The program will skip this process.")
    #             else:
    #                 try:
    #                     process.kill()
    #                 except psutil.NoSuchProcess:
    #                     if print_detail:
    #                         print(f"进程{pName}（{pid}）已被提前结束。\nProcess {pName} ({pid}) has been terminated in advance.")
    #                 else:
    #                     if print_detail:
    #                         print(f"已结束进程{pName}（{pid}）。\nProcess {pName} ({pid}) terminated.")

#-----------------------------------------------------------------------------
# Main
#-----------------------------------------------------------------------------

connector.start()
