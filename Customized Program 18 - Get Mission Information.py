from lcu_driver import Connector
import json, os, pandas, re, shutil, time, unicodedata
from wcwidth import wcswidth
from openpyxl import load_workbook

#=============================================================================
# * 声明（Declaration）
#=============================================================================
# 作者（Author）：       XHXIAIEIN
# 更新（Last update）：  2021/01/08
# 主页（Home page）：    https://github.com/XHXIAIEIN/LeagueCustomLobby/
#=============================================================================

#-----------------------------------------------------------------------------
# 工具库（Tool library）
#-----------------------------------------------------------------------------
#  - lcu-driver 
#    https://github.com/sousa-andre/lcu-driver
#-----------------------------------------------------------------------------

connector = Connector()

async def get_summoner_data(connection):
    data = await connection.request('GET', '/lol-summoner/v1/current-summoner')
    summoner = await data.json()
    print("displayName:    %s" %(summoner["gameName"] + "#" + summoner["tagLine"]))
    print("summonerId:     %s" %(summoner["summonerId"]))
    print("puuid:          %s" %(summoner["puuid"]))
    print("-")


#-----------------------------------------------------------------------------
#  lockfile
#-----------------------------------------------------------------------------
async def update_lockfile(connection):
    path = os.path.join(connection.installation_path.encode('gb18030').decode('utf-8'), 'lockfile')
    if os.path.isfile(path):
        file = open(path, 'w+')
        text = "LeagueClient:%d:%d:%s:%s" %(connection.pid, connection.port, connection.auth_key, connection.protocols[0])
        file.write(text)
        file.close()
    return None

async def get_lockfile(connection):
    path = os.path.join(connection.installation_path.encode('gb18030').decode('utf-8'), 'lockfile')
    if os.path.isfile(path):
        file = open(path, 'r')
        text = file.readline().split(':')
        file.close()
        print(connection.address)
        print(f'riot    {connection.auth_key}')
        return connection.auth_key
    return None

#-----------------------------------------------------------------------------
# 获取英雄联盟中的所有游戏类型信息（Get all game types' information in League of Legends）
#-----------------------------------------------------------------------------
def count_nonASCII(s: str): #统计一个字符串中占用命令行2个宽度单位的字符个数（Count the number of characters that take up 2 width unit in CMD）
    return sum([unicodedata.east_asian_width(character) in ("F", "W") for character in list(str(s))])

def format_df(df: pandas.DataFrame, width_exceed_ask: bool = True, direct_print: bool = False, print_header: bool = True, print_index: bool = False, reserve_index = False, start_index = 0, header_align: str = "^", align: str = "^", align_replicate_rule: str = "all"): #按照每列最长字符串的命令行宽度加上2，再根据每个数据的中文字符数量决定最终格式化输出的字符串宽度（Get the width of the longest string of each column, add it by 2, and substract it by the number of each cell string's Chinese characters to get the final width for each cell to print using `format` function）
    old_index = df.index
    df.index = range(start_index, len(df) + start_index)
    maxLens = {}
    maxWidth = shutil.get_terminal_size()[0]
    fields = df.columns.tolist()
    for field in fields:
        maxLens[field] = max(0 if len(df) == 0 else max(map(lambda x: wcswidth(str(x)), df[field])), wcswidth(str(field))) + 2
    index_len = max(len(str(start_index)), len(str(start_index + len(df) - 1)))
    if sum(maxLens.values()) + 2 * (len(fields) - 1) > maxWidth or print_index and index_len + sum(maxLens.values()) + 2 * len(fields) > maxWidth:
        if width_exceed_ask:
            print("单行数据字符串输出宽度超过当前终端窗口宽度！是否继续？（输入任意键继续，否则直接打印该数据框。）\nThe output width of each record string exceeds the current width of the terminal window! Continue? (Input anything to continue, or null to directly print this dataframe.)")
            if input() == "":
                #print(df)
                result = str(df)
                return (result, maxLens)
        elif direct_print:
            print("单行数据字符串输出宽度超过当前终端窗口宽度！将直接打印该数据框！\nThe output width of each record string exceeds the current width of the terminal window! The program is going to directly print this dataframe!")
            result = str(df)
            return (result, maxLens)
        else:
            print("单行数据字符串输出宽度超过当前终端窗口宽度！将继续格式化输出！\nThe output width of each record string exceeds the current width of the terminal window! The program is going on formatted printing!")
    result = ""
    #确定各列的排列方向（Determine the alignments of all columns）
    if isinstance(header_align, str) and isinstance(align, str):
        if not all(map(lambda x: x in {"<", "^", ">"}, header_align)) or not all(map(lambda x: x in {"<", "^", ">"}, align)):
            print('排列方式字符串参数错误！排列方式必须是“<”“^”或者“>”中的一个。请修改排列方式字符串参数。\nParameter ERROR of the alignment string! The alignment value must be one of {"<", "^", ">"}. Please change the alignment string parameter.')
        if len(header_align) == 0: #指定为空字符串，即默认居中输出（Specifying it as a null string means output centered by default）
            header_alignments = ["^"] * df.shape[1]
        elif len(header_align) == 1:
            header_alignments = [header_align] * df.shape[1]
        else:
            header_alignments_tmp = list(header_align)
            if len(header_align) < df.shape[1]:
                if align_replicate_rule == "last":
                    header_alignments = header_alignments_tmp + [header_alignments_tmp[-1]] * len(df.shape[1] - len(header_align))
                else:
                    if align_replicate_rule != "all":
                        print("排列方式列表补充规则不合法！将默认采用全部填充。\nAlignment list supplement rule illegal! The whole alignment string will be replicated.")
                    header_alignments = header_alignments_tmp * (df.shape[1] // len(header_align)) + header_alignments_tmp[:df.shape[1] % len(header_align)]
            else:
                header_alignments = header_alignments_tmp[:df.shape[1]]
        if len(align) == 0:
            alignments = ["^"] * df.shape[1]
        elif len(align) == 1:
            alignments = [align] * df.shape[1]
        else:
            alignments_tmp = list(align)
            if len(align) < df.shape[1]:
                if align_replicate_rule == "last":
                    alignments = alignments_tmp + [alignments_tmp[-1]] * len(df.shape[1] - len(align))
                else:
                    if align_replicate_rule != "all":
                        print("排列方式列表补充规则不合法！将默认采用全部填充。\nAlignment list supplement rule illegal! The whole alignment string will be replicated.")
                    alignments = alignments_tmp * (df.shape[1] // len(align)) + alignments_tmp[:df.shape[1] % len(align)]
            else:
                alignments = alignments_tmp[:df.shape[1]]
        if print_header:
            if print_index:
                result += " " * (index_len + 2)
            for i in range(df.shape[1]):
                field = fields[i]
                tmp = "{0:{align}{w}}".format(field, align = header_alignments[i], w = maxLens[str(field)] - count_nonASCII(str(field)))
                result += tmp
                #print(tmp, end = "")
                if i != df.shape[1] - 1:
                    result += "  "
                    #print("  ", end = "")
            result += "\n"
            #print()
        index = start_index
        for i in range(df.shape[0]):
            if print_index:
                result += "{0:>{w}}".format(old_index[index - start_index] if reserve_index else index, w = index_len) + "  "
            for j in range(df.shape[1]):
                field = fields[j]
                cell = str(list(df[field])[i])
                tmp = "{0:{align}{w}}".format(cell, align = alignments[j], w = maxLens[field] - count_nonASCII(cell))
                result += tmp
                #print(tmp, end = "")
                if j != df.shape[1] - 1:
                    result += "  "
                    #print("  ", end = "")
            if i != df.shape[0] - 1:
                result += "\n"
            #print() #注意这里的缩进和上一行不同（Note that here the indentation is different from the last line）
            index += 1
    else:
        print("排列方式参数错误！请传入字符串。\nAlignment parameter ERROR! Please pass a string instead.")
    return (result, maxLens)

def get_info_name(info: dict, mode = 1) -> str:
    if not isinstance(info, dict) or not all(i in info for i in ["displayName", "gameName", "tagLine"]):
        print("您的召唤师信息格式有误！\nERROR format of summoner information!")
        name = ""
        exit()
    else:
        if info["displayName"] or info["gameName"]:
            if info["gameName"] and info["tagLine"]:
                name = info["gameName"] + "#" + info["tagLine"]
            elif not info["tagLine"] and info["gameName"]:
                name = info["gameName"]
            else:
                name = info["displayName"]
        else: #新玩家属于这种类型（This case matches new players）
            if mode == 1:
                name = str(info["puuid"])
            elif mode == 2: #仅用于设置召唤师数据保存路径（Designed to set the summoner name directory）
                name = "0. 新玩家\\" + str(info["puuid"])
            elif mode == 3: #仅用于设置召唤师数据保存路径（Designed to set the summoner name directory）
                name = "0. New Player\\" + str(info["puuid"])
    return name

async def get_mission_info(connection):
    #设置输出路径（Set the output directory）
    info = await (await connection.request("GET", "/lol-summoner/v1/current-summoner")).json()
    displayName = get_info_name(info)
    platform_config = await (await connection.request("GET", "/lol-platform-config/v1/namespaces")).json()
    platformId = platform_config["LoginDataPacket"]["platformId"]
    platform_TENCENT = {"BGP1": "全网通区 男爵领域（Baron Zone）", "BGP2": "峡谷之巅（Super Zone）", "EDU1": "教育网专区（CRENET Server）", "HN1": "电信一区 艾欧尼亚（Ionia）", "HN2": "电信二区 祖安（Zaun）", "HN3": "电信三区 诺克萨斯（Noxus 1）", "HN4": "电信四区 班德尔城（Bandle City）", "HN4_NEW": "电信四区 班德尔城（Bandle City）", "HN5": "电信五区 皮尔特沃夫（Piltover）", "HN6": "电信六区 战争学院（the Institute of War）", "HN7": "电信七区 巨神峰（Mount Targon）", "HN8": "电信八区 雷瑟守备（Noxus 2）", "HN9": "电信九区 裁决之地（the Proving Grounds）", "HN10": "电信十区 黑色玫瑰（the Black Rose）", "HN11": "电信十一区 暗影岛（Shadow Isles）", "HN12": "电信十二区 钢铁烈阳（the Iron Solari）", "HN13": "电信十三区 水晶之痕（Crystal Scar）", "HN14": "电信十四区 均衡教派（the Kinkou Order）", "HN15": "电信十五区 影流（the Shadow Order）", "HN16": "电信十六区 守望之海（Guardian's Sea）", "HN17": "电信十七区 征服之海（Conqueror's Sea）", "HN18": "电信十八区 卡拉曼达（Kalamanda）", "HN19": "电信十九区 皮城警备（Piltover Wardens）", "PBE": "体验服 试炼之地（Chinese PBE）", "WT1": "网通一区 比尔吉沃特（Bilgewater）", "WT1_NEW": "网通一区 比尔吉沃特（Bilgewater）", "WT2": "网通二区 德玛西亚（Demacia）", "WT2_NEW": "网通二区 德玛西亚（Demacia）", "WT3": "网通三区 弗雷尔卓德（Freljord）", "WT3_NEW": "网通三区 弗雷尔卓德（Freljord）", "WT4": "网通四区 无畏先锋（House Crownguard）", "WT4_NEW": "网通四区 无畏先锋（House Crownguard）", "WT5": "网通五区 恕瑞玛（Shurima）", "WT6": "网通六区 扭曲丛林（Twisted Treeline）", "WT7": "网通七区 巨龙之巢（the Dragon Camp）", "FORCES": "比赛服 艾欧尼亚（Tournament - Ionia）", "NJ100": "联盟一区", "GZ100": "联盟二区", "CQ100": "联盟三区", "TJ100": "联盟四区", "TJ101": "联盟五区", "PREPBE": "试炼之地 临时过渡服务器（Chinese PBE Temporary）"}
    platform_RIOT = {"ME1": "中东服（Middle East）", "BR1": "巴西服（Brazil）", "EUN1": "北欧和东欧服（Europe Nordic & East）", "EUW1": "西欧服（Europe West）", "JP1": "日服（Japan）", "KR": "韩服（Republic of Korea）", "LA1": "北拉美服（Latin America North）", "LA2": "南拉美服（Latin America South）", "NA1": "北美服（North America）", "OC1": "大洋洲服（Oceania）", "TR1": "土耳其服（Turkey）", "RU": "俄罗斯服（Russia）", "PH2": "菲律宾服（Philippines）", "SG2": "新加坡服（Singapore）", "TH2": "泰服（Thailand）", "TW2": "台服（Taiwan, Hong Kong and Macau）", "VN2": "越南服（Vietnam）", "PBE1": "测试服（Public Beta Environment）"}
    platform_GARENA = {"PH1": "菲律宾服（Philippines）", "SG1": "新加坡服（Singapore, Malaysia and Indonesia）", "TW1": "台服（Taiwan, Hong Kong and Macau）", "VN1": "越南服（Vietnam）", "TH1": "泰服（Thailand）"}
    platform = {"TENCENT": "国服（TENCENT）", "RIOT": "外服（RIOT）", "GARENA": "竞舞（GARENA）"}
    riot_client_info = await (await connection.request("GET", "/riotclient/command-line-args")).json()
    client_info = {}
    for i in range(len(riot_client_info)):
        try:
            client_info[riot_client_info[i].split("=")[0]] = riot_client_info[i].split("=")[1]
        except IndexError:
            pass
    region = client_info["--region"]
    if region == "TENCENT":
        platform_folder = "召唤师信息（Summoner Information）\\" + "国服（TENCENT）" + "\\" + platform_TENCENT[platformId]
        folder = platform_folder + "\\" + get_info_name(info, 2)
    elif region == "GARENA":
        platform_folder = "召唤师信息（Summoner Information）\\" + "竞舞（GARENA）" + "\\" + platform_GARENA[platformId]
        folder = platform_folder + "\\" + get_info_name(info, 2)
    else: #拳头公司与竞舞娱乐公司的合同于2023年1月终止（In January 2023, Riot Games ended its contract with Garena）
        platform_folder = "召唤师信息（Summoner Information）\\" + "外服（RIOT）" + "\\" + (platform_RIOT | platform_GARENA)[platformId]
        folder = platform_folder + "\\" + get_info_name(info, 3)
    platform_config_filepath = platform_folder + "\\" + "platform_config_namespaces.json"
    while True:
        try:
            with open(platform_config_filepath, "w", encoding = "utf-8") as fp:
                json.dump(platform_config, fp, indent = 4, ensure_ascii = False)
        except FileNotFoundError: #这里需要注意是否具有创建文件夹的权限。下同（Pay attention to the authority to create the folder. So are the following）
            os.makedirs(os.path.dirname(platform_config_filepath), exist_ok = True)
        else:
            break
    #定义常量字典（Define constant dictionaries）
    celebrationTypes = {"NONE": "无", "TOAST": "干杯", "VIGNETTE": "花饰", "VIGNETTE_LARGE_REWARDS_ONLY": "高等奖励专用花饰", "VIGNETTE_REWARDS_ONLY": "奖励专用花饰"}
    clientNotifyLevels = {"ALWAYS": "总是", "NONE": "从不"}
    displayTypes = {"AFTER_COMPLETION": "完成后显示", "ALWAYS": "总是显示", "CELEBRATION_ONLY": "仅在庆祝时显示", "TUTORIAL_ONLY": "仅新手教程显示"}
    missionTypes = {"ONETIME": "一次性", "REPEATING": "可重复"}
    metadataMissionTypes = {"": "", "always": "永久"}
    objectiveStatus_dict = {"DUMMY": "占位", "ELIGIBLE": "具备资格"}
    objectiveTypes = {"": "", "CHAMPION_MASTERY": "英雄成就", "EOGDATA": "对局数据", "INGEST": "新手学习", "LEGS": "英雄联盟传统玩法", "SERIES_COMPLETION": "任务系列", "TFT_ELIMINATION": "云顶之弈淘汰任务"}
    rewardGroupStrategies = {"": "", "ALL_GROUPS": "所有分组", "SELECT_GROUPS": "选定分组"}
    rewardTypes = {"BLUE_ESSENCE": "蓝色精萃", "CHAMPION_SHARD": "英雄碎片", "CHAMPION_SKIN_SHARD": "皮肤碎片", "CHAMPION_TOKEN": "永久英雄", "CLIENT_FEATURE": "游戏模式", "GAME_QUEUE": "游戏队列", "HEXTECH_CHEST": "海克斯科技宝箱", "HEXTECH_KEY": "海克斯科技钥匙", "HEXTECH_KEY_SHARD": "海克斯科技钥匙碎片", "MISSION_PROGRESS": "其它任务完成进度", "ORANGE_ESSENCE": "橙色精粹", "PROGRESSION": "通行证进度", "REWARD_GROUP": "多重奖励", "RIOT_POINTS": "点券", "SPELL_BOOK_PAGE": "符文页", "SUMMONER_ICON": "召唤师图标", "SUMMONER_SPELL": "召唤师技能", "WARD_SKIN_SHARD": "守卫皮肤碎片", "XP": "召唤师等级经验值"}
    missionStatus_dict = {"COMPLETED": "已完成", "DUMMY": "用于测试", "PENDING": "未完成", "SELECT_REWARDS": "选择奖励", "UPCOMING": "未激活"}
    gameTypes = {"lol": "英雄联盟", "tft": "云顶之弈"}
    objectivesTypes = {"kNonPooledObjectives": "非池化目标", "kPooledObjectives": "池化目标"}
    #获取数据资源（Get data resources）
    missions = await (await connection.request("GET", "/lol-missions/v1/missions")).json()
    lolObjectives = await (await connection.request("GET", "/lol-objectives/v1/objectives/lol")).json() #该列表中的两个字典的“objectives”键的值相同（Values of the "objectives" key of both dictionaries in this list are the same）
    lolObjectives_active = []
    if len(lolObjectives) > 1:
        current_time = time.time() * 1000
        for objective in lolObjectives:
            if not (objective["startDate"] < current_time - time.localtime().tm_gmtoff < objective["endDate"]):
                lolObjectives_active.append(objective)
    lolObjective = lolObjectives[0] if len(lolObjectives_active) == 0 else lolObjectives_active[0]
    tftObjectives = await (await connection.request("GET", "/lol-objectives/v1/objectives/tft")).json()
    tftObjectives_active = []
    if len(tftObjectives) > 1:
        current_time = time.time() * 1000
        for objective in tftObjectives:
            if not (objective["startDate"] < current_time - time.localtime().tm_gmtoff < objective["endDate"]):
                tftObjectives_active.append(objective)
    tftObjective = tftObjectives[0] if len(tftObjectives_active) == 0 else tftObjectives_active[0]
    objectives = [lolObjective, tftObjective]
    mission_objectiveGroup_map = {}
    objectiveGroupId_order = {} #用于后续数据框排序——第一关键字（Used as the first keyword for the subsequent dataframe sorting）
    weight = 0
    for objective in objectives:
        for objectiveGroup in objective["objectives"]:
            objectiveGroupId_order[objectiveGroup["id"]] = weight
            weight += 1
            for mission in objectiveGroup["missions"]:
                mission_objectiveGroup_map[(mission["id"], mission["sequence"])] = {"id": objectiveGroup["id"], "localizedTag": objectiveGroup["localizedTag"], "localizedTitle": objectiveGroup["localizedTitle"]}
    #整理数据（Sort out data）
    mission_header = {"backgroundImageUrl": "任务背景图片链接", "celebrationType": "庆祝类型", "clientNotifyLevel": "客户端通知类型", "completedDate": "任务完成时间戳", "completionExpression": "任务完成方法", "cooldownTimeMillis": "任务刷新间隔时间（毫秒）", "description": "任务描述", "displayType": "任务显示类型", "earnedDate": "任务获取时间戳", "endTime": "任务结束时间戳", "expiringWarnings": "过期警告", "helperText": "任务附加说明", "iconImageUrl": "任务图标链接", "id": "任务序号", "internalName": "任务内置名", "isNew": "新任务", "lastUpdatedTimestamp": "上次更新任务时间戳", "locale": "语言", "media": "媒体信息", "missionLineText": "任务标题文本", "missionType": "任务类型", "requirements": "任务激活要求", "rewards": "任务奖励详细信息", "sequence": "任务序列号", "seriesName": "任务系列名称", "startTime": "任务开始时间戳", "status": "任务状态", "title": "任务标题", "viewed": "已查看", "completedTime": "任务完成时间", "earnedTime": "任务获取时间", "endDateTime": "任务结束时间", "lastUpdatedTime": "上次更新任务时间", "startDateTime": "任务开始时间", "rewardDescriptions": "任务奖励描述", "display attributes": "任务显示属性", "display locations": "任务显示位置", "metadata chain": "元数据：羁绊", "metadata chainSize": "元数据：羁绊任务数量", "metadata missionType": "元数据：任务类型", "metadata order": "元数据：顺序", "metadata weekNum": "元数据：周次", "metadata xpReward": "元数据：经验值奖励", "metadata npeRewardPack index": "元数据：新玩家奖励序号", "metadata npeRewardPack majorReward": "元数据：新玩家奖励：主要奖励", "metadata npeRewardPack minorRewards": "元数据：新玩家奖励：次要奖励", "metadata npeRewardPack premiumReward": "元数据：新玩家奖励：高级奖励已激活", "metadata npeRewardPack rewardKey": "元数据：新玩家奖励：奖励代码", "metadata tutorial displayRewards": "元数据：新手教程：显示奖励", "metadata tutorial queueId": "元数据：新手教程：队列序号", "metadata tutorial stepNumber": "元数据：新手教程：步骤", "metadata tutorial useChosenChampion": "元数据：新手教程：使用给定英雄", "metadata tutorial useQuickSearchMatchmaking": "元数据：新手教程：使用快速模式匹配系统", "rewardStrategy groupStrategy": "任务奖励分组", "rewardStrategy selectMaxGroupCount": "任务奖励分组序号上限", "rewardStrategy selectMinGroupCount": "任务奖励分组序号下限", "objective description": "目标描述", "objective hasObjectiveBasedReward": "目标特定奖励", "objective requirements": "目标激活要求", "objective rewardGroups": "目标奖励分组", "objective sequence": "目标序列号", "objective status": "目标状态", "objective type": "目标类型", "objective progress currentProgress": "目标当前进度", "objective progress lastViewedProgress": "目标上次查看时进度", "objective progress totalCount": "目标完成所需进度", "objectiveGroup id": "目标分组序号", "objectiveGroup localizedTag": "目标分组标签", "objectiveGroup localizedTitle": "目标分组标题"}
    mission_header_keys = list(mission_header.keys())
    mission_data = {}
    for i in range(len(mission_header_keys)):
        key = mission_header_keys[i]
        mission_data[key] = []
    for mission in missions:
        for objective in mission["objectives"]:
            for i in range(len(mission_header_keys)):
                key = mission_header_keys[i]
                if i <= 34:
                    if i == 1: #庆祝类型（`celebrationType`）
                        mission_data[key].append(celebrationTypes[mission[key]])
                    elif i == 2: #客户端通知类型（`clientNotifyLevel`）
                        mission_data[key].append(clientNotifyLevels[mission[key]])
                    elif i == 7: #显示类型（`displayType`）
                        mission_data[key].append(displayTypes[mission[key]])
                    elif i == 20: #任务类型（`missionTypes`）
                        mission_data[key].append(missionTypes[mission[key]])
                    elif i == 26: #状态（`status`）
                        mission_data[key].append(missionStatus_dict[mission[key]])
                    elif i >= 29 and i <= 33: #时间类键（Time-type keys）
                        subkey_dict = {29: "completedDate", 30: "earnedDate", 31: "endTime", 32: "lastUpdatedTimestamp", 33: "startTime"}
                        try:
                            timeStr = "" if mission[subkey_dict[i]] == -1 else time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(mission[subkey_dict[i]] // 1000))
                        except OSError:
                            timeStr = mission[subkey_dict[i]]
                        mission_data[key].append(timeStr)
                    elif i == 34: #任务奖励描述（`rewardDescriptions`）
                        rewardDescriptions = []
                        for reward in mission["rewards"]:
                            if not reward["description"] in rewardDescriptions: #同一项奖励描述可能被拆成多项奖励类型不同的奖励（One description can be distributed into multiple rewards with different rewardTypes）
                                rewardDescriptions.append(reward["description"])
                        mission_data[key].append(rewardDescriptions)
                    else:
                        mission_data[key].append(mission[key])
                elif i <= 55: #该代码框架适用于纯嵌套字典类型的值（This code frame applies to pure nested dictionary-type values）
                    tmpObj_ptr = mission
                    for subkey_iter in key.split():
                        tmpObj_ptr = tmpObj_ptr[subkey_iter]
                    if i == 39: #元数据：任务类型（`metadata missionType`）
                        mission_data[key].append(metadataMissionTypes[tmpObj_ptr])
                    elif i == 53: #奖励分组（`rewardStrategy groupStrategy`）
                        mission_data[key].append(rewardGroupStrategies[tmpObj_ptr])
                    else:
                        mission_data[key].append(tmpObj_ptr)
                elif i <= 65:
                    tmpObj_ptr = objective
                    for j in range(1, len(key.split())):
                        tmpObj_ptr = tmpObj_ptr[key.split()[j]]
                    if i == 61: #目标状态（`objective status`）
                        mission_data[key].append(objectiveStatus_dict[tmpObj_ptr])
                    elif i == 62: #目标类型（`objective type`）
                        mission_data[key].append(objectiveTypes[tmpObj_ptr])
                    else:
                        mission_data[key].append(tmpObj_ptr)
                else:
                    subkey = key.split()[1]
                    if (mission["id"], mission["sequence"]) in mission_objectiveGroup_map:
                        mission_data[key].append(mission_objectiveGroup_map[(mission["id"], mission["sequence"])][subkey])
                    else:
                        mission_data[key].append("")
    objective_group_header = {"backgroundImage": "目标背景图片链接", "endDate": "目标结束时间戳", "gameType": "游戏类型", "isActive": "可用性", "objectivesCategories": "目标分类", "startDate": "目标开始时间戳", "uuid": "通用唯一识别码", "endTime": "目标结束时间", "startTime": "目标开始时间", "objectiveGroup backgroundImage": "目标分组背景图片链接", "objectiveGroup endDate": "目标分组结束时间戳", "objectiveGroup id": "目标分组序号", "objectiveGroup isEnabled": "目标分组可用性", "objectiveGroup isPooledMission": "目标分组池化任务", "objectiveGroup localizedTag": "目标分组标签", "objectiveGroup localizedTitle": "目标分组标题", "objectiveGroup maxRefresh": "目标分组最大刷新次数", "objectiveGroup objectivesType": "目标分组类型", "objectiveGroup priority": "目标分组排列顺序", "objectiveGroup refreshInterval": "目标分组刷新间隔（分钟）", "objectiveGroup startDate": "目标分组开始时间戳", "objectiveGroup tag": "目标分组标签", "objectiveGroup endTime": "目标分组结束时间", "objectiveGroup startTime": "目标分组开始时间"}
    objective_group_header_keys = list(objective_group_header.keys())
    objective_group_data = {}
    for i in range(len(objective_group_header_keys)):
        key = objective_group_header_keys[i]
        objective_group_data[key] = []
    for objective in objectives:
        for objectiveGroup in objective["objectives"]:
            for i in range(len(objective_group_header_keys)):
                key = objective_group_header_keys[i]
                if i <= 8:
                    if i == 2: #游戏类型（`gameType`）
                        objective_group_data[key].append(gameTypes[objective[key]])
                    elif i >= 7: #时间类键（Time-type keys）
                        subkey_dict = {7: "endDate", 8: "startDate"}
                        timeStr = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(objective[subkey_dict[i]] // 1000))
                        objective_group_data[key].append(timeStr)
                    else:
                        objective_group_data[key].append(objective[key])
                else:
                    if i >= 22: #时间类键（Time-type keys）
                        subkey_dict = {22: "endDate", 23: "startDate"}
                        timeStr = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(objectiveGroup[subkey_dict[i]] // 1000))
                        objective_group_data[key].append(timeStr)
                    else:
                        subkey = key.split()[1]
                        objective_group_data[key].append(objectiveGroup[subkey])
    #数据框列序整理（Dataframe column ordering）
    mission_statistics_output_order = [68, 67, 66, 13, 27, 14, 6, 19, 11, 24, 23, 21, 35, 36, 56, 60, 58, 62, 64, 63, 65, 61, 57, 59, 26, 34, 22, 53, 54, 55, 33, 31, 32, 30, 4, 29, 5, 20, 7, 1, 2, 15, 28, 18, 10, 0, 12, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52]
    mission_data_organized = {}
    for i in mission_statistics_output_order:
        key = mission_header_keys[i]
        mission_data_organized[key] = mission_data[key]
    mission_df = pandas.DataFrame(data = mission_data_organized)
    for column in mission_df:
        if mission_df[column].dtype == "bool":
            mission_df[column] = mission_df[column].astype(str)
            for i in range(len(mission_df)):
                mission_df.loc[i, column] = "√" if mission_df[column][i] == "True" else ""
    mission_df = pandas.concat([pandas.DataFrame([mission_header])[mission_df.columns], mission_df], ignore_index = True)
    objective_group_statistics_output_order = [2, 6, 3, 4, 7, 8, 15, 14, 21, 11, 18, 17, 12, 13, 23, 22, 19, 16, 9, 0]
    objective_group_data_organized = {}
    for i in objective_group_statistics_output_order:
        key = objective_group_header_keys[i]
        objective_group_data_organized[key] = objective_group_data[key]
    objective_group_df = pandas.DataFrame(data = objective_group_data_organized)
    for column in objective_group_df:
        if objective_group_df[column].dtype == "bool":
            objective_group_df[column] = objective_group_df[column].astype(str)
            for i in range(len(objective_group_df)):
                objective_group_df.loc[i, column] = "√" if objective_group_df[column][i] == "True" else ""
    objective_group_df = pandas.concat([pandas.DataFrame([objective_group_header])[objective_group_df.columns], objective_group_df], ignore_index = True)
    #数据框排序（Dataframe sorting）
    mission_df_data = mission_df.loc[1:, :]
    mission_df_data_sorted = mission_df_data.sort_values(by = "objectiveGroup id", key = lambda x: x.map(objectiveGroupId_order), ascending = True)
    mission_df_sorted = pandas.concat([mission_df.iloc[:1, :], mission_df_data_sorted])
    #保存文件（Save the files）
    excel_name = f"Player Mission - {displayName}.xlsx" 
    excel_name_sorted = f"Player Mission - {displayName} (sorted).xlsx"
    workbook_exist = True
    while True:
        try:
            with pandas.ExcelWriter(path = os.path.join(folder, excel_name), mode = "a", if_sheet_exists = "replace") as writer:
                currentTime = time.strftime("%Y-%m-%d", time.localtime(time.time()))
                mission_df_sorted.to_excel(excel_writer = writer, sheet_name = "Missions - " + currentTime + "_" + platformId)
                objective_group_df.to_excel(excel_writer = writer, sheet_name = "Objectives - " + currentTime + "_" + platformId)
            print('玩家目标和任务信息已保存为“%s”！\nPlayer objective and mission information is saved as "%s"!' %(os.path.join(folder, excel_name), os.path.join(folder, excel_name)))
        except PermissionError:
            print("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
            input()
        except FileNotFoundError:
            workbook_exist = False
            os.makedirs(folder, exist_ok = True)
            with pandas.ExcelWriter(path = os.path.join(folder, excel_name)) as writer:
                currentTime = time.strftime("%Y-%m-%d", time.localtime(time.time()))
                mission_df_sorted.to_excel(excel_writer = writer, sheet_name = "Missions - " + currentTime + "_" + platformId)
                objective_group_df.to_excel(excel_writer = writer, sheet_name = "Objectives - " + currentTime + "_" + platformId)
            print('玩家目标和任务信息已保存为“%s”！\nPlayer objective and mission information is saved as "%s"!' %(os.path.join(folder, excel_name), os.path.join(folder, excel_name)))
            break
        else:
            break
    #工作表排序（Worksheet ordering）
    if workbook_exist:
        print("警告：由于该文件已存在，本次导出已追加新工作表到工作簿的末尾。这可能导致工作表顺序的错乱。是否需要对工作表进行排序？（输入任意键排序，否则不排序）\nWarning: Because the excel workbook has existed, new sheets are appended to the last of the original sheet list. This may result in the disarrangement of worksheet order. Do you want to sort the sheets? (Input anything to sort the sheets, or null to skip sorting)")
        sort = bool(input())
        if sort:
            mission_loaded = True
            print("正在读取刚刚创建的工作表……\nLoading the workbook just created ...")
            while True:
                try:
                    wb = load_workbook(os.path.join(folder, excel_name))
                except FileNotFoundError:
                    print('任务信息工作簿读取失败！请确保“%s”文件夹内含有名为“%s”的工作簿。如果需要退出程序，请输入“0”。\nERROR reading the Missions workbook! Please make sure the workbook "%s" is in the folder "%s". If you want to exit the program, please submit "0".' %(folder, excel_name, excel_name, folder))
                    mission_reload = input()
                    if mission_reload == "0":
                        mission_loaded = False
                        break
                else:
                    break
            if mission_loaded:
                sheetnames = wb.sheetnames #第一次获取原工作簿的工作表名称列表（The first time to get the sheet name list of the original workbook）
                print("请选择排序方式：\nPlease select an ordering pattern:\n1\t时间优先（默认）【Time in priority (by default)】\n2\t类别优先（Type in priority）")
                op = input()
                print("正在创建顺序工作表列表……\nCreating the ordered sheet list ...")
                date_re = re.compile(r"\d{4}-\d{2}-\d{2}") #设置正则表达式识别
                if op == "" or op[0] != "2": #按照时间优先的原则对工作表进行排序，时间相同则任务工作表在前，目标工作表在后（Sort the sheets by time in priority. If the times are the same, then the mission sheet is arranged in front of the objective sheet）
                    sheetname_date_list = list(map(lambda x: date_re.search(x).group(), sheetnames)) #从工作表名称提取日期信息形成列表（Extract the dates from the sheetnames to form a list）
                    sheetname_type_list = list(map(lambda x: x.split()[0], sheetnames)) #从工作表名称提取数据类型信息形成列表（Extract the data types from the sheetnames to form a list）
                    sheetname_platform_list = list(map(lambda x: x.split("_")[1], sheetnames)) #从工作表名称提取大区信息形成列表（Extract the platformId from the sheetnames to form a list）
                    sheetname_tmpDf = pandas.DataFrame(data = [sheetnames, sheetname_date_list, sheetname_type_list, sheetname_platform_list]).stack().unstack(0) #创建一个四列数据框，各列分别是完整工作表名、日期信息、数据类型信息和大区信息（Create a 4-column dataframe whose columns are the complete sheetname, date, data type and platformId）
                    sheetnames_sorted = sheetname_tmpDf.sort_values(by = [1, 2, 3], ascending = [True, True, True]).iloc[:, 0].tolist() #将工作表名按照第一关键字——日期信息正序排列，第二关键字——数据类型信息正序排列（先任务后目标），第三关键字——大区信息正序排列（Order the sheetnames according to the ascending order of the first keyword - date, the ascending order of the second keyword - data type and the ascending order of the third keyword - platformId）
                else:
                    sheets_Missions = [sheet_iter for sheet_iter in sheetnames if sheet_iter.startswith("Missions")] #提取任务类型的工作表名称（Extract the names of the sheets containing mission data）
                    sheets_Objectives = [sheet_iter for sheet_iter in sheetnames if sheet_iter.startswith("Objectives")] #提取目标类型的工作表名称（Extract the names of the sheets containing objective data）
                    sheets_Missions = sorted(sheets_Missions, key = lambda x: date_re.search(x).group()) #按照日期正序排列任务类型的工作表名称（Order the mission sheetnames according to the ascending order of dates）
                    sheets_Objectives = sorted(sheets_Objectives, key = lambda x: date_re.search(x).group()) #按照日期正序排列目标类型的工作表名称（Order the objective sheetnames according to the ascending order of dates）
                    sheetnames_sorted = sheets_Missions + sheets_Objectives #合并列表得到先按类别排列、再按日期排列的工作表名称（Combine the lists to get the sheetname list ordered firstly by data type and secondly by date）
                #下面排列所有工作表（The following code arrange all sheets）
                print("正在排序……\nOrdering ...")
                for i in range(len(sheetnames_sorted)): #排序的思路是每次将一个工作表根据其在原工作表列表中的索引和在顺序工作表列表中的索引的差值进行移动（The main idea of sheets' sorting is to move each sheet according to the difference of the indices between in the original sheet list and in the ordered sheet list）
                    sheetnames = wb.sheetnames #因为一次移动可能导致很多其它工作表的位置发生变化，所以必须每次都重新获取工作表列表（Because a moving event may result in location change of many other sheets, the sheet list must be obtained each time）
                    sheetname_iter = sheetnames_sorted[i] #这里以顺序工作表为迭代器进行遍历，因为顺序工作表是固定不变的（Here the ordered sheet list acts as the iterator to be traversed, for the ordered sheet list is fixed）
                    if sheetnames[i] != sheetname_iter:
                        preIndex = sheetnames.index(sheetname_iter)
                        wb.move_sheet(sheetname_iter, i - preIndex) #注意移动距离数应当是排序后的索引减去排序前的索引（Note that the moving offset should be the index in the ordered list subtracted by that in the original list）
                    #print("排序进度（Ordering process）：%d/%d\t工作表名称（Sheet name）： %s" %(i + 1, len(sheetnames_sorted), sheetname_iter))
                print('正在保存中……\nSaving the ordered workbook ...')
                wb.save(os.path.join(folder, excel_name_sorted))
                print('排序完成！排好序的工作簿已保存为“%s”。\nOrdering finished! The ordered workbook is saved as "%s".\n' %(excel_name_sorted, excel_name_sorted))

async def check_repeating_missions(connection):
    #查看可重复任务的刷新状态（Check repeating missions' cooldown status）
    while True:
        missions = await (await connection.request("GET", "/lol-missions/v1/missions")).json()
        repeating_missions = [mission for mission in missions if mission["cooldownTimeMillis"] != -1]
        completed_repeating_missions = [mission for mission in repeating_missions if mission["status"] == "COMPLETED"]
        pending_repeating_missions = [mission for mission in repeating_missions if mission["status"] == "PENDING"]
        if len(repeating_missions) != 0:
            print("在该服务器上检测到可重复任务。\nDetected repeating missions on this server.")
            if len(completed_repeating_missions) != 0:
                print("以下任务已完成：\nThe following missions are completed:")
                for mission in completed_repeating_missions:
                    rewardDescriptions = []
                    for reward in mission["rewards"]:
                        if not reward["description"] in rewardDescriptions:
                            rewardDescriptions.append(reward["description"]) #奖励（Reward）
                    completedTime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(mission["completedDate"] // 1000)) #完成时间（Completed time）
                    cooldown = mission["cooldownTimeMillis"] / 1000
                    cooldown_hour = cooldown // 3600
                    cooldown_minute = cooldown // 3600 % 60
                    cooldown_second = cooldown % 60
                    cooldown_str = "%d:%02d:%02d" %(cooldown_hour, cooldown_minute, cooldown_second) #刷新间隔（Cooldown）
                    refreshDate = mission["completedDate"] + mission["cooldownTimeMillis"]
                    refreshTime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(refreshDate // 1000)) #刷新时间（Refresh time）
                    cooldown_remaining = refreshDate // 1000 - time.time()
                    cooldown_remaining_hour = cooldown_remaining // 3600
                    cooldown_remaining_minute = cooldown_remaining % 3600 // 60
                    cooldown_remaining_second = cooldown_remaining % 60
                    cooldown_remaining_str = "%d:%02d:%02d" %(cooldown_remaining_hour, cooldown_remaining_minute, cooldown_remaining_second) #剩余时间（Cooldown remaining）
                    missionData_simple = {"项目": ["标题", "序号", "序列号", "描述", "奖励", "完成时间", "刷新间隔", "刷新时间", "剩余时间"], "Items": ["title", "id", "sequence", "description", "rewards", "completedTime", "cooldown", "refreshTime", "cooldownRemaining"], "值": [mission["title"], mission["id"], mission["sequence"], mission["description"], rewardDescriptions, completedTime, cooldown_str, refreshTime, cooldown_remaining_str]}
                    #print("标题（Title）： %s\n序号（Id）： %s\n序列号（Sequence）： %d\n描述（Description）： %s\n奖励（Reward）： %s\n完成时间（Completed time）： %s\n刷新间隔（Cooldown）： %s\n刷新时间（Refresh time）： %s\n剩余时间：（Cooldown remaining）： %s\n" %(mission["title"], mission["id"], mission["sequence"], mission["description"], rewardDescriptions, completedTime, cooldown_str, refreshTime, cooldown_remaining_str))
                    missionDf_simple = pandas.DataFrame(data = missionData_simple)
                    print(format_df(missionDf_simple, print_index = True, header_align = "^", align = "^^>")[0], end = "\n\n")
            if len(pending_repeating_missions) != 0:
                print("\n以下任务等待完成：\nThe following missions are pending:")
                for mission in pending_repeating_missions:
                    rewardDescriptions = []
                    for reward in mission["rewards"]:
                        if not reward["description"] in rewardDescriptions:
                            rewardDescriptions.append(reward["description"]) #奖励（Reward）
                    missionData_simple = {"项目": ["标题", "序号", "序列号", "描述", "奖励"], "Items": ["title", "id", "sequence", "description", "rewards"], "值": [mission["title"], mission["id"], mission["sequence"], mission["description"], rewardDescriptions]}
                    #print("标题（Title）： %s\n序号（Id）： %s\n序列号（Sequence）： %d\n描述（Description）： %s\n奖励（Reward）： %s\n" %(mission["title"], mission["id"], mission["sequence"], mission["description"], rewardDescriptions))
                    missionDf_simple = pandas.DataFrame(data = missionData_simple)
                    print(format_df(missionDf_simple, print_index = True, header_align = "^", align = "^^>")[0], end = "\n\n")
            print("\n是否更新任务状态？（输入任意键以更新，否则退出程序。）\nUpdate the missions? (Submit any non-empty string to update, or null to exit the program.)")
            if not bool(input()):
                break
        else:
            break

#-----------------------------------------------------------------------------
# websocket
#-----------------------------------------------------------------------------
@connector.ready
async def connect(connection):
    await get_summoner_data(connection)
    print("是否导出所有任务信息？（输入任意键不导出，否则导出。）\nDo you want to export all missions' information? (Submit any non-empty string to refuse exporting, or null to export.)")
    if not bool(input()):
        await get_mission_info(connection)
    await check_repeating_missions(connection)

#-----------------------------------------------------------------------------
# Main
#-----------------------------------------------------------------------------

connector.start()
